"""Общие функции пакетных экспериментов маршрутов (Excel, точки, сводки).

Используется ``route_variants_experiment``, ``heat_weather_experiment``, ``route_batch_experiment``
и ``run_variants_over_weather_cases`` (см. ``_run_variants_weather_batch``).
Выходные .xlsx пишутся в ``bike_router/experiment_outputs/`` (см. ``experiment_output_xlsx_path``).
"""

from __future__ import annotations

import hashlib
import json
import logging
import math
import os
import re
import sys
import time
import uuid
from pathlib import Path
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, time as dt_time, timedelta, timezone
from math import asin, cos, radians, sin, sqrt
from typing import Any, Dict, Iterator, List, Optional, Sequence, Tuple

try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None  # type: ignore[misc, assignment]

_log = logging.getLogger(__name__)

from bike_router.services.routing_criteria import CRITERION_KEYS

_CRITERIA_ROUTE_MEAN_COLS = tuple(f"route_mean_{k}" for k in CRITERION_KEYS)

# Периодический INFO в батче (шаг = одна пара O–D × один сценарий погоды, все варианты).
DEFAULT_BATCH_LOG_EVERY = 100

_NOISY_BATCH_LOGGER_NAMES = (
    "bike_router.engine",
    "bike_router.services.weather",
    "bike_router.services.routing",
    "bike_router.services.graph",
    "bike_router.services.area_graph_cache",
    "bike_router.services.corridor_graph_cache",
    "urllib3",
    "requests",
)


def _set_quiet_mode_for_batch(verbose: bool) -> None:
    """Тихий режим: INFO движка/OSM/HTTP — только WARNING и выше."""
    if verbose:
        return
    for name in _NOISY_BATCH_LOGGER_NAMES:
        logging.getLogger(name).setLevel(logging.WARNING)


# Порядок как в engine._UNIFIED_ROUTE_ORDER
EXPECTED_VARIANTS: Tuple[str, ...] = (
    "shortest",
    "full",
    "green",
    "heat",
    "stress",
    "heat_stress",
)

PROFILES: Tuple[str, ...] = ("cyclist", "pedestrian")

# Перебор буфера коридора в этом эксперименте; дальше — пропуск пары (см. RouteNotFoundError).
EXPERIMENT_CORRIDOR_EXPAND_M: Tuple[float, ...] = (10.0, 100.0)

# Часовой пояс для «локального часа» в multi-day Open-Meteo (можно переопределить BATCH_WEATHER_TZ).
_DEFAULT_BATCH_WEATHER_TZ = "Europe/Samara"

# Фиксированные параметры батча (CLI урезан).
DEFAULT_ROUTE_BATCH_SEED = 42
DEFAULT_MIN_SPACING_M = 300.0
DEFAULT_MAX_SAMPLE_ATTEMPTS = 50_000

# Все Excel-выходы пакетных экспериментов — в этой папке (от корня пакета bike_router).
EXPERIMENT_OUTPUT_DIR_NAME = "experiment_outputs"


def experiment_output_xlsx_path(*, script_stem: str) -> str:
    """Путь к .xlsx: ``bike_router/experiment_outputs/{script_stem}_YYYYMMDD_HHMMSS.xlsx`` (UTC)."""
    pkg_root = Path(__file__).resolve().parent.parent
    out_dir = pkg_root / EXPERIMENT_OUTPUT_DIR_NAME
    out_dir.mkdir(parents=True, exist_ok=True)
    fn = datetime.now(timezone.utc).strftime(f"{script_stem}_%Y%m%d_%H%M%S.xlsx")
    return str(out_dir / fn)
# Режим ``past``: столько дней архива, каждый день — 14:00 локально.
PAST_ARCHIVE_DAYS = 10
PAST_ARCHIVE_LOCAL_HOUR = 14

# Режим ``--weather test``: фиксированный ISO без Open-Meteo.
SYNTHETIC_TEST_WEATHER_ISO = "2000-06-15T12:00:00+00:00"


@dataclass(frozen=True)
class SyntheticWeatherCase:
    case_id: str
    temperature_c: float
    precipitation_mm: float
    wind_speed_ms: float
    wind_gusts_ms: float
    cloud_cover_pct: float
    humidity_pct: float
    shortwave_radiation_wm2: float
    # ISO для сезонного профиля (зима / апрель / лето); по умолчанию — летний synthetic.
    weather_time_iso: str = SYNTHETIC_TEST_WEATHER_ISO
    snowfall_cm_h: float = 0.0
    snow_depth_m: float = 0.0
    weather_code: Optional[int] = None
    # Метео ° — откуда дует (Open-Meteo); None — без direction-aware ветра.
    wind_direction_deg: Optional[float] = None


def weather_windows_test_grid() -> List[SyntheticWeatherCase]:
    """3×2×2×3 = 36 синтетических сценариев (температура × дождь × ветер × облачность)."""
    temps = (0.0, 12.5, 25.0)
    rains = ((False, 0.0), (True, 1.5))
    winds = ((False, 2.0, 3.0), (True, 9.0, 14.0))
    clouds = ((0.0, 750.0), (50.0, 400.0), (100.0, 80.0))
    out: List[SyntheticWeatherCase] = []
    for T in temps:
        for rain_on, precip in rains:
            for wind_flag, w_ms, g_ms in winds:
                for c_pct, sw in clouds:
                    hum = 85.0 if rain_on else 55.0
                    tid = int(round(T * 10))
                    cid = f"test_T{tid}_R{int(rain_on)}_W{int(wind_flag)}_C{int(c_pct)}"
                    out.append(
                        SyntheticWeatherCase(
                            case_id=cid,
                            temperature_c=T,
                            precipitation_mm=precip,
                            wind_speed_ms=w_ms,
                            wind_gusts_ms=g_ms,
                            cloud_cover_pct=c_pct,
                            humidity_pct=hum,
                            shortwave_radiation_wm2=sw,
                        )
                    )
    assert len(out) == 36
    return out


def weather_summer_test_grid_with_wind_dirs() -> List[SyntheticWeatherCase]:
    """36 летних × 4 направления = 144 (полный обход углов; для батча см. ``weather_summer_heat_grid``)."""
    base = weather_windows_test_grid()
    dirs = (0.0, 90.0, 180.0, 270.0)
    out: List[SyntheticWeatherCase] = []
    for c in base:
        for d in dirs:
            out.append(
                SyntheticWeatherCase(
                    case_id=f"{c.case_id}_WD{int(d)}",
                    temperature_c=c.temperature_c,
                    precipitation_mm=c.precipitation_mm,
                    wind_speed_ms=c.wind_speed_ms,
                    wind_gusts_ms=c.wind_gusts_ms,
                    cloud_cover_pct=c.cloud_cover_pct,
                    humidity_pct=c.humidity_pct,
                    shortwave_radiation_wm2=c.shortwave_radiation_wm2,
                    weather_time_iso=c.weather_time_iso,
                    snowfall_cm_h=c.snowfall_cm_h,
                    snow_depth_m=c.snow_depth_m,
                    weather_code=c.weather_code,
                    wind_direction_deg=float(d),
                )
            )
    assert len(out) == 144
    return out


def weather_summer_heat_grid() -> List[SyntheticWeatherCase]:
    """90 летних synthetic: слабый ветер — одно направление (0°), сильный — 4 (0/90/180/270°).

    3×2×3 базовых по темп/дождь/облачность × (9 calm + 18 strong wind rows) = 90.
    """
    temps = (0.0, 12.5, 25.0)
    rains = ((False, 0.0), (True, 1.5))
    winds = ((False, 2.0, 3.0), (True, 9.0, 14.0))
    clouds = ((0.0, 750.0), (50.0, 400.0), (100.0, 80.0))
    out: List[SyntheticWeatherCase] = []
    for T in temps:
        for rain_on, precip in rains:
            for wind_flag, w_ms, g_ms in winds:
                wind_dirs = (0.0,) if not wind_flag else (0.0, 90.0, 180.0, 270.0)
                for c_pct, sw in clouds:
                    hum = 85.0 if rain_on else 55.0
                    tid = int(round(T * 10))
                    cid0 = f"test_T{tid}_R{int(rain_on)}_W{int(wind_flag)}_C{int(c_pct)}"
                    for d in wind_dirs:
                        out.append(
                            SyntheticWeatherCase(
                                case_id=f"{cid0}_WD{int(d)}",
                                temperature_c=T,
                                precipitation_mm=precip,
                                wind_speed_ms=w_ms,
                                wind_gusts_ms=g_ms,
                                cloud_cover_pct=c_pct,
                                humidity_pct=hum,
                                shortwave_radiation_wm2=sw,
                                weather_time_iso=SYNTHETIC_TEST_WEATHER_ISO,
                                wind_direction_deg=float(d),
                            )
                        )
    assert len(out) == 90
    return out


# Зима: для сезонной модели и снега (UTC февраль).
WINTER_SYNTH_WEATHER_ISO = "2000-02-15T12:00:00+00:00"
# Ранняя весна: слабая зелень до середины апреля.
EARLY_SPRING_SYNTH_WEATHER_ISO = "2000-04-05T12:00:00+00:00"


def weather_winter_synthetic_grid() -> List[SyntheticWeatherCase]:
    """3×3×3×2 = 54: температура × снегопад × глубина снега на земле × ветер (зимние synthetic-кейсы)."""
    temps = (-20.0, -10.0, 0.0)
    fresh_levels = (
        (0.0, "F0"),
        (0.45, "F1"),
        (3.2, "F2"),
    )
    depths_m = (0.0, 0.10, 0.22)
    winds = ((2.0, 3.0, "W0"), (11.0, 15.0, "W1"))
    out: List[SyntheticWeatherCase] = []
    for T in temps:
        for sf, ftag in fresh_levels:
            for dm in depths_m:
                for w_ms, g_ms, wtag in winds:
                    tid = int(round(T))
                    cid = f"winter_T{tid}_{ftag}_D{int(round(dm * 1000))}_{wtag}"
                    hum = 72.0 if sf > 0.1 else 58.0
                    sw = 120.0 if T < 0 else 280.0
                    out.append(
                        SyntheticWeatherCase(
                            case_id=cid,
                            temperature_c=T,
                            precipitation_mm=0.0,
                            wind_speed_ms=w_ms,
                            wind_gusts_ms=g_ms,
                            cloud_cover_pct=85.0 if sf > 0.2 else 55.0,
                            humidity_pct=hum,
                            shortwave_radiation_wm2=sw,
                            weather_time_iso=WINTER_SYNTH_WEATHER_ISO,
                            snowfall_cm_h=sf,
                            snow_depth_m=dm,
                            weather_code=73 if sf > 1.0 else (71 if sf > 0.1 else None),
                        )
                    )
    assert len(out) == 54
    return out


def weather_winter_synthetic_grid_with_wind_dirs() -> List[SyntheticWeatherCase]:
    """54 зимних × 4 направления = 216 (полный обход; для батча см. ``weather_winter_heat_grid``)."""
    base = weather_winter_synthetic_grid()
    dirs = (0.0, 90.0, 180.0, 270.0)
    out: List[SyntheticWeatherCase] = []
    for c in base:
        for d in dirs:
            out.append(
                SyntheticWeatherCase(
                    case_id=f"{c.case_id}_WD{int(d)}",
                    temperature_c=c.temperature_c,
                    precipitation_mm=c.precipitation_mm,
                    wind_speed_ms=c.wind_speed_ms,
                    wind_gusts_ms=c.wind_gusts_ms,
                    cloud_cover_pct=c.cloud_cover_pct,
                    humidity_pct=c.humidity_pct,
                    shortwave_radiation_wm2=c.shortwave_radiation_wm2,
                    weather_time_iso=c.weather_time_iso,
                    snowfall_cm_h=c.snowfall_cm_h,
                    snow_depth_m=c.snow_depth_m,
                    weather_code=c.weather_code,
                    wind_direction_deg=float(d),
                )
            )
    assert len(out) == 216
    return out


def weather_winter_heat_grid() -> List[SyntheticWeatherCase]:
    """135 зимних synthetic: W0 (слабый ветер) — одно направление; W1 — четыре.

    27 пространственных комбинаций × (1 + 4) направлений по силе ветра = 135.
    """
    base = weather_winter_synthetic_grid()
    out: List[SyntheticWeatherCase] = []
    for c in base:
        strong = float(c.wind_speed_ms) >= 9.0
        wind_dirs = (0.0, 90.0, 180.0, 270.0) if strong else (0.0,)
        for d in wind_dirs:
            out.append(
                SyntheticWeatherCase(
                    case_id=f"{c.case_id}_WD{int(d)}",
                    temperature_c=c.temperature_c,
                    precipitation_mm=c.precipitation_mm,
                    wind_speed_ms=c.wind_speed_ms,
                    wind_gusts_ms=c.wind_gusts_ms,
                    cloud_cover_pct=c.cloud_cover_pct,
                    humidity_pct=c.humidity_pct,
                    shortwave_radiation_wm2=c.shortwave_radiation_wm2,
                    weather_time_iso=c.weather_time_iso,
                    snowfall_cm_h=c.snowfall_cm_h,
                    snow_depth_m=c.snow_depth_m,
                    weather_code=c.weather_code,
                    wind_direction_deg=float(d),
                )
            )
    assert len(out) == 135
    return out


def _batch_profiles_from_arg(arg: str) -> Tuple[str, ...]:
    v = (arg or "both").strip().lower()
    if v == "both":
        return ("cyclist", "pedestrian")
    if v == "cyclist":
        return ("cyclist",)
    if v == "pedestrian":
        return ("pedestrian",)
    raise ValueError("profiles: ожидается both | cyclist | pedestrian")


def _batch_output_xlsx_path() -> str:
    """Уникальное имя: route_experiment_batch_YYYYMMDD_HHMMSS.xlsx (UTC)."""
    return datetime.now(timezone.utc).strftime(
        "route_experiment_batch_%Y%m%d_%H%M%S.xlsx"
    )


def centroid_lat_lon_for_weather(poly: Any) -> Tuple[float, float]:
    """Центр bounds полигона: (lat, lon) для одного запроса погоды."""
    minx, miny, maxx, maxy = poly.bounds
    lon_mid = (float(minx) + float(maxx)) * 0.5
    lat_mid = (float(miny) + float(maxy)) * 0.5
    return lat_mid, lon_mid


def kwargs_fixed_snapshot_from_case(
    case: SyntheticWeatherCase,
    *,
    weather_time_iso: Optional[str] = None,
) -> Dict[str, Any]:
    wt = weather_time_iso or getattr(
        case, "weather_time_iso", SYNTHETIC_TEST_WEATHER_ISO
    )
    return {
        "weather_mode": "fixed-snapshot",
        "use_live_weather": False,
        "weather_time": wt,
        "temperature_c": case.temperature_c,
        "precipitation_mm": case.precipitation_mm,
        "wind_speed_ms": case.wind_speed_ms,
        "cloud_cover_pct": case.cloud_cover_pct,
        "humidity_pct": case.humidity_pct,
        "wind_gusts_ms": case.wind_gusts_ms,
        "shortwave_radiation_wm2": case.shortwave_radiation_wm2,
        "snowfall_cm_h": float(getattr(case, "snowfall_cm_h", 0.0) or 0.0),
        "snow_depth_m": float(getattr(case, "snow_depth_m", 0.0) or 0.0),
        "weather_code": getattr(case, "weather_code", None),
        "wind_direction_deg": getattr(case, "wind_direction_deg", None),
    }


def resolve_live_weather_once_for_polygon(
    poly: Any,
    *,
    settings: Any,
    departure_iso: Optional[str] = None,
) -> Tuple[Any, str, Any, str, float, float]:
    """Один вызов Open-Meteo в центре полигона; возвращает snap, src, wp, dep, lat, lon."""
    from bike_router.engine import _resolve_route_weather

    lat_c, lon_c = centroid_lat_lon_for_weather(poly)
    start = (lat_c, lon_c)
    end = (lat_c, lon_c)
    dep = departure_iso or datetime.now(timezone.utc).replace(
        microsecond=0
    ).isoformat()
    _thermal = {
        "hot_tree": float(settings.heat_hot_tree_bonus_scale),
        "hot_open": float(settings.heat_hot_open_sky_penalty_scale),
        "cold_canyon": float(settings.heat_cold_building_canyon_bonus_scale),
        "cold_tree_damp": float(settings.heat_cold_tree_bonus_damping),
        "response": float(settings.heat_weather_response_scale),
    }
    w_snap, wsrc, wp = _resolve_route_weather(
        start,
        end,
        request_mode="auto",
        use_live_weather=True,
        weather_time=None,
        departure_time=dep,
        temperature_c=None,
        precipitation_mm=None,
        wind_speed_ms=None,
        cloud_cover_pct=None,
        humidity_pct=None,
        wind_gusts_ms=None,
        wind_direction_deg=None,
        shortwave_radiation_wm2=None,
        thermal_scales=_thermal,
        settings=settings,
    )
    return w_snap, wsrc, wp, dep, lat_c, lon_c


def meta_append_batch_weather_snapshot(
    base: List[Tuple[str, Any]],
    *,
    snap: Any,
    source: str,
    departure_iso: str,
    center_lat: float,
    center_lon: float,
    experiment_kind: str,
    wp: Any = None,
) -> List[Tuple[str, Any]]:
    """Поля снимка погоды для листа «Метаданные» (сырые величины).

    Если передан ``wp`` (WeatherWeightParams после build), добавляются сезон и сила snow-модели.
    """
    out = list(base)
    out.append(("experiment_kind", experiment_kind))
    out.append(("batch_weather_center_lat", center_lat))
    out.append(("batch_weather_center_lon", center_lon))
    out.append(("batch_weather_departure_iso", departure_iso))
    out.append(("snapshot_temperature_c", getattr(snap, "temperature_c", None)))
    out.append(
        (
            "snapshot_apparent_temperature_c",
            getattr(snap, "apparent_temperature_c", None),
        )
    )
    out.append(("snapshot_precipitation_mm", getattr(snap, "precipitation_mm", None)))
    out.append(
        (
            "snapshot_precipitation_probability",
            getattr(snap, "precipitation_probability", None),
        )
    )
    out.append(("snapshot_wind_speed_ms", getattr(snap, "wind_speed_ms", None)))
    out.append(("snapshot_wind_gusts_ms", getattr(snap, "wind_gusts_ms", None)))
    out.append(
        ("snapshot_wind_direction_deg", getattr(snap, "wind_direction_deg", None))
    )
    out.append(("snapshot_cloud_cover_pct", getattr(snap, "cloud_cover_pct", None)))
    out.append(("snapshot_humidity_pct", getattr(snap, "humidity_pct", None)))
    out.append(
        (
            "snapshot_shortwave_radiation_wm2",
            getattr(snap, "shortwave_radiation_wm2", None),
        )
    )
    out.append(
        ("snapshot_snowfall_cm_h", getattr(snap, "snowfall_cm_h", None)),
    )
    out.append(("snapshot_snow_depth_m", getattr(snap, "snow_depth_m", None)))
    out.append(("snapshot_weather_code", getattr(snap, "weather_code", None)))
    out.append(("weather_snapshot_source", source))
    if wp is not None and bool(getattr(wp, "enabled", False)):
        out.append(
            ("snapshot_routing_season", str(getattr(wp, "routing_season", "") or ""))
        )
        out.append(
            (
                "snapshot_routing_season_calendar",
                str(getattr(wp, "routing_season_calendar", "") or ""),
            )
        )
        out.append(
            (
                "snapshot_routing_season_source",
                str(getattr(wp, "routing_season_source", "") or ""),
            )
        )
        out.append(
            (
                "snapshot_season_green_mult",
                float(getattr(wp, "season_green_route_mult", 1.0) or 1.0),
            )
        )
        out.append(
            (
                "snapshot_snow_model_strength",
                float(getattr(wp, "snow_model_strength", 0.0) or 0.0),
            )
        )
    return out


def _zoneinfo_or_raise(name: str) -> Any:
    if ZoneInfo is None:
        raise RuntimeError(
            "Нужен Python 3.9+ с tzdata для часового пояса погоды (zoneinfo)."
        )
    return ZoneInfo(name)


def weather_windows_past_local_days(
    *,
    past_days: int,
    local_hour: int,
    tz_name: str,
) -> List[Tuple[str, str]]:
    """Календарные даты в прошлом (вчера, позавчера, …) и ISO-время для Open-Meteo.

    Возвращает список из ``past_days`` пар
    ``(weather_date YYYY-MM-DD, weather_time локальный ISO)``.
    """
    tz = _zoneinfo_or_raise(tz_name)
    if not (0 <= int(local_hour) <= 23):
        raise ValueError("local_hour должен быть 0..23")
    out: List[Tuple[str, str]] = []
    today = datetime.now(tz).date()
    for k in range(1, int(past_days) + 1):
        d = today - timedelta(days=k)
        dt = datetime.combine(d, dt_time(hour=int(local_hour)), tzinfo=tz)
        out.append((d.isoformat(), dt.isoformat()))
    return out


SUMMARY_NUM_KEYS = (
    "length_m",
    "time_s",
    "climb_m",
    "descent_m",
    "max_gradient_pct",
    "green_percent",
    "avg_trees_pct",
    "avg_grass_pct",
    "stress_cost_total",
    "weather_temperature_c",
    "weather_apparent_temperature_c",
    "weather_precipitation_mm",
    "weather_precipitation_probability",
    "weather_wind_speed_ms",
    "weather_wind_gusts_ms",
    "weather_wind_direction_deg",
    "weather_wind_direction_aware",
    "weather_cloud_cover_pct",
    "weather_humidity_pct",
    "weather_shortwave_radiation_wm2",
    "weather_snowfall_cm_h",
    "weather_snow_depth_m",
    "weather_weather_code",
    "weather_routing_season",
    "weather_routing_season_calendar",
    "weather_routing_season_source",
    "weather_season_green_mult",
    "weather_season_tree_heat_mult",
    "weather_season_stress_route_mult",
    "weather_season_stairs_route_mult",
    "weather_season_wind_orientation_route_mult",
    "weather_stress_route_regime_factor",
    "weather_snow_model_strength",
    "weather_snow_export_phys_amp",
    "weather_snow_export_stress_amp",
    "weather_snow_export_surface_amp",
    "weather_heat_continuous",
    "heat_tree_shade_bonus",
    "heat_open_sky_penalty",
    "heat_building_shade_bonus",
    "heat_covered_bonus",
    "heat_wind_open_penalty",
    "heat_wet_surface_penalty",
    "heat_norm_temp",
    "heat_norm_rain",
    "heat_norm_wind",
    "heat_norm_gust",
    "heat_norm_cloud",
    "heat_norm_humidity",
    "heat_norm_cold_like",
    "heat_norm_snow_depth",
    "heat_norm_snow_fresh",
    "route_open_sky_share",
    "route_building_shade_share",
    "route_covered_share",
    "route_bad_wet_surface_share",
    "route_winter_harsh_surface_share",
    "route_wind_direction_aware",
    "route_mean_wind_to_street_angle_deg",
    "route_mean_heat_directional_wind_exp",
    "route_mean_heat_building_wind_factor",
    "route_frac_wind_along_open_hostile",
    "route_frac_wind_cross_building_screen",
    *_CRITERIA_ROUTE_MEAN_COLS,
    "stress_route_regime_factor",
    "route_stairs_length_m",
    "route_stairs_length_fraction",
    "route_winter_open_stress_proxy",
    "weather_test_temperature_c",
    "weather_test_precipitation_mm",
    "weather_test_wind_speed_ms",
    "weather_test_wind_gusts_ms",
    "weather_test_wind_direction_deg",
    "weather_test_cloud_cover_pct",
    "weather_test_humidity_pct",
    "weather_test_shortwave_radiation_wm2",
)


def _weather_metrics_from_route(r: Any) -> Dict[str, Any]:
    """Числовые поля снимка Open-Meteo из ответа движка (для Excel и сводок)."""
    empty: Dict[str, Any] = {
        "weather_temperature_c": None,
        "weather_apparent_temperature_c": None,
        "weather_precipitation_mm": None,
        "weather_precipitation_probability": None,
        "weather_wind_speed_ms": None,
        "weather_wind_gusts_ms": None,
        "weather_wind_direction_deg": None,
        "weather_wind_direction_aware": None,
        "weather_cloud_cover_pct": None,
        "weather_humidity_pct": None,
        "weather_shortwave_radiation_wm2": None,
        "weather_snowfall_cm_h": None,
        "weather_snow_depth_m": None,
        "weather_weather_code": None,
        "weather_routing_season": None,
        "weather_routing_season_calendar": None,
        "weather_routing_season_source": None,
        "weather_season_green_mult": None,
        "weather_season_tree_heat_mult": None,
        "weather_season_stress_route_mult": None,
        "weather_season_stairs_route_mult": None,
        "weather_season_wind_orientation_route_mult": None,
        "weather_stress_route_regime_factor": None,
        "weather_snow_model_strength": None,
        "weather_snow_export_phys_amp": None,
        "weather_snow_export_stress_amp": None,
        "weather_snow_export_surface_amp": None,
        "weather_heat_continuous": None,
        "heat_tree_shade_bonus": None,
        "heat_open_sky_penalty": None,
        "heat_building_shade_bonus": None,
        "heat_covered_bonus": None,
        "heat_wind_open_penalty": None,
        "heat_wet_surface_penalty": None,
        "heat_norm_temp": None,
        "heat_norm_rain": None,
        "heat_norm_wind": None,
        "heat_norm_gust": None,
        "heat_norm_cloud": None,
        "heat_norm_humidity": None,
        "heat_norm_cold_like": None,
        "heat_norm_snow_depth": None,
        "heat_norm_snow_fresh": None,
    }
    w = getattr(r, "weather", None)
    if not w or not bool(getattr(w, "enabled", False)):
        return empty
    s = getattr(w, "snapshot", None)
    if s is None:
        return empty

    def _f(name: str) -> Optional[float]:
        v = getattr(s, name, None)
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    out = {
        "weather_temperature_c": _f("temperature_c"),
        "weather_apparent_temperature_c": _f("apparent_temperature_c"),
        "weather_precipitation_mm": _f("precipitation_mm"),
        "weather_precipitation_probability": _f("precipitation_probability"),
        "weather_wind_speed_ms": _f("wind_speed_ms"),
        "weather_wind_gusts_ms": _f("wind_gusts_ms"),
        "weather_wind_direction_deg": _f("wind_direction_deg"),
        "weather_wind_direction_aware": bool(
            getattr(w, "wind_direction_available", False)
        ),
        "weather_cloud_cover_pct": _f("cloud_cover_pct"),
        "weather_humidity_pct": _f("humidity_pct"),
        "weather_shortwave_radiation_wm2": _f("shortwave_radiation_wm2"),
        "weather_snowfall_cm_h": _f("snowfall_cm_h"),
        "weather_snow_depth_m": _f("snow_depth_m"),
        "weather_weather_code": int(getattr(s, "weather_code"))
        if getattr(s, "weather_code", None) is not None
        else None,
        "weather_routing_season": str(getattr(w, "routing_season", "") or ""),
        "weather_routing_season_calendar": str(
            getattr(w, "routing_season_calendar", "") or ""
        ),
        "weather_routing_season_source": str(
            getattr(w, "routing_season_source", "") or ""
        ),
        "weather_season_green_mult": float(
            getattr(w, "season_green_route_mult", 1.0) or 1.0
        ),
        "weather_season_tree_heat_mult": float(
            getattr(w, "season_tree_heat_route_mult", 1.0) or 1.0
        ),
        "weather_season_stress_route_mult": float(
            getattr(w, "season_stress_route_mult", 1.0) or 1.0
        ),
        "weather_season_stairs_route_mult": float(
            getattr(w, "season_stairs_route_mult", 1.0) or 1.0
        ),
        "weather_season_wind_orientation_route_mult": float(
            getattr(w, "season_wind_orientation_route_mult", 1.0) or 1.0
        ),
        "weather_stress_route_regime_factor": float(
            getattr(w, "stress_route_regime_factor", 1.0) or 1.0
        ),
        "weather_snow_model_strength": float(
            getattr(w, "snow_model_strength", 0.0) or 0.0
        ),
        "weather_snow_export_phys_amp": float(
            getattr(w, "snow_export_phys_amp", 1.0) or 1.0
        ),
        "weather_snow_export_stress_amp": float(
            getattr(w, "snow_export_stress_amp", 1.0) or 1.0
        ),
        "weather_snow_export_surface_amp": float(
            getattr(w, "snow_export_surface_amp", 1.0) or 1.0
        ),
        "weather_heat_continuous": bool(getattr(w, "heat_continuous", False)),
        "heat_tree_shade_bonus": None,
        "heat_open_sky_penalty": None,
        "heat_building_shade_bonus": None,
        "heat_covered_bonus": None,
        "heat_wind_open_penalty": None,
        "heat_wet_surface_penalty": None,
        "heat_norm_temp": None,
        "heat_norm_rain": None,
        "heat_norm_wind": None,
        "heat_norm_gust": None,
        "heat_norm_cloud": None,
        "heat_norm_humidity": None,
        "heat_norm_cold_like": None,
    }
    hm = getattr(w, "heat_microclimate", None) or {}
    if isinstance(hm, dict):
        mapping = {
            "tree_shade_bonus": "heat_tree_shade_bonus",
            "open_sky_penalty": "heat_open_sky_penalty",
            "building_shade_bonus": "heat_building_shade_bonus",
            "covered_bonus": "heat_covered_bonus",
            "wind_open_penalty": "heat_wind_open_penalty",
            "wet_surface_penalty": "heat_wet_surface_penalty",
            "norm_temp_norm": "heat_norm_temp",
            "norm_rain_norm": "heat_norm_rain",
            "norm_wind_norm": "heat_norm_wind",
            "norm_gust_norm": "heat_norm_gust",
            "norm_cloud_norm": "heat_norm_cloud",
            "norm_humidity_norm": "heat_norm_humidity",
            "norm_cold_like_norm": "heat_norm_cold_like",
            "norm_snow_depth_norm": "heat_norm_snow_depth",
            "norm_snow_fresh_norm": "heat_norm_snow_fresh",
        }
        for sk, dk in mapping.items():
            if sk in hm:
                try:
                    out[dk] = float(hm[sk])
                except (TypeError, ValueError):
                    pass
    return out


def _ensure_pkg_path() -> None:
    root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    if root not in sys.path:
        sys.path.insert(0, root)


def haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Расстояние по поверхности сферы (м)."""
    r = 6371000.0
    p1, p2 = radians(lat1), radians(lat2)
    dphi = radians(lat2 - lat1)
    dl = radians(lon2 - lon1)
    a = sin(dphi / 2) ** 2 + cos(p1) * cos(p2) * sin(dl / 2) ** 2
    return 2 * r * asin(min(1.0, sqrt(a)))


def _point_id_fmt(i: int) -> str:
    return f"P{i + 1:02d}"


@dataclass
class SampledPoint:
    idx: int
    lat: float
    lon: float
    nearest_node_id: int


def _iter_directed_pairs(n: int) -> Iterator[Tuple[int, int]]:
    for i in range(n):
        for j in range(n):
            if i == j:
                continue
            yield i, j


def _iter_undirected_pairs(n: int) -> Iterator[Tuple[int, int]]:
    """Пары i<j — один раз на неориентированную пару точек (для heat synthetic)."""
    for i in range(n):
        for j in range(i + 1, n):
            yield i, j


def _direction_key(a: int, b: int) -> str:
    return f"{_point_id_fmt(a)}->{_point_id_fmt(b)}"


def sample_points_in_polygon(
    *,
    poly: Any,
    n_target: int,
    rng: Any,
    engine: Any,
    min_spacing_m: float,
    max_attempts: int,
) -> List[SampledPoint]:
    """Случайные точки в полигоне: уникальность, min_spacing_m, валидация на графе."""
    from shapely.geometry import Point

    import osmnx as ox

    from bike_router.exceptions import PointOutsideZoneError

    minx, miny, maxx, maxy = poly.bounds
    out: List[SampledPoint] = []
    attempts = 0
    while len(out) < n_target and attempts < max_attempts:
        attempts += 1
        x = float(rng.uniform(minx, maxx))
        y = float(rng.uniform(miny, maxy))
        p = Point(x, y)
        if not poly.contains(p):
            continue
        lat, lon = y, x
        if any(
            haversine_m(lat, lon, sp.lat, sp.lon) < min_spacing_m for sp in out
        ):
            continue
        try:
            engine._validate_point((lat, lon), "sample")  # noqa: SLF001
        except PointOutsideZoneError:
            continue
        G = engine.graph
        if G is None:
            raise RuntimeError("Граф не загружен после warmup")
        nid = int(ox.distance.nearest_nodes(G, X=lon, Y=lat))
        out.append(SampledPoint(idx=len(out), lat=lat, lon=lon, nearest_node_id=nid))

    if len(out) < n_target:
        raise RuntimeError(
            f"Не удалось набрать {n_target} точек за {max_attempts} попыток "
            f"(получено {len(out)}). Увеличьте max_attempts или ослабьте min_spacing_m."
        )
    _log.info(
        "Точки: готово %d шт. (попыток %d, min %.0f м)",
        len(out),
        attempts,
        min_spacing_m,
    )
    return out


def _warnings_text(r: Any) -> Tuple[int, str]:
    qh = r.quality_hints
    if not qh or not qh.warnings:
        return 0, ""
    w = qh.warnings
    return len(w), " | ".join(str(x) for x in w)


def _stress_fields(r: Any) -> Dict[str, Any]:
    hs = r.heat_stress
    out: Dict[str, Any] = {
        "stress_cost_total": float(r.stress_cost_total or 0.0),
        "avg_stress_lts": None,
        "max_stress_lts": None,
        "high_stress_segments_count": int(r.high_stress_segments_count or 0),
        "stressful_intersections_count": int(r.stressful_intersections_count or 0),
    }
    if hs:
        out["avg_stress_lts"] = float(hs.avg_stress_lts)
        out["max_stress_lts"] = float(hs.max_stress_lts)
        if out["stress_cost_total"] == 0.0 and hs.stress_cost_total:
            out["stress_cost_total"] = float(hs.stress_cost_total)
        if not out["high_stress_segments_count"] and hs.high_stress_segments_count:
            out["high_stress_segments_count"] = int(hs.high_stress_segments_count)
        if (
            not out["stressful_intersections_count"]
            and hs.stressful_intersections_count
        ):
            out["stressful_intersections_count"] = int(
                hs.stressful_intersections_count
            )
    return out


def route_to_raw_row(
    *,
    experiment_id: str,
    seed: int,
    route_id: int,
    profile: str,
    origin_id: int,
    dest_id: int,
    o_lat: float,
    o_lon: float,
    d_lat: float,
    d_lon: float,
    r: Any,
    baseline_full: Any = None,
    weather_date: str = "",
    test_weather_meta: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    st = _stress_fields(r)
    wn, wt = _warnings_text(r)
    elev = r.elevation
    green = r.green
    wc = r.weather.weather_time if r.weather else None
    wsrc = str(r.weather.source or "") if r.weather else ""
    geom = r.geometry or []
    geom_json = json.dumps(geom, ensure_ascii=False, separators=(",", ":"))

    out: Dict[str, Any] = {
        "route_id": route_id,
        "experiment_id": experiment_id,
        "seed": seed,
        "profile": profile,
        "variant_key": r.mode,
        "variant_label": r.variant_label or "",
        "origin_point_id": _point_id_fmt(origin_id),
        "destination_point_id": _point_id_fmt(dest_id),
        "direction_key": _direction_key(origin_id, dest_id),
        "direction_order": "forward"
        if origin_id < dest_id
        else "reverse",
        "origin_lat": o_lat,
        "origin_lon": o_lon,
        "destination_lat": d_lat,
        "destination_lon": d_lon,
        "route_built_at_utc": r.route_built_at_utc or "",
        "weather_date": weather_date or (str(wc)[:10] if wc else ""),
        "weather_time": wc or "",
        "weather_source": wsrc,
        **_weather_metrics_from_route(r),
        "length_m": float(r.length_m),
        "length_km": round(float(r.length_m) / 1000.0, 6),
        "time_s": float(r.time_s),
        "time_min": round(float(r.time_s) / 60.0, 4),
        "time_display": r.time_display or "",
        "climb_m": float(elev.climb_m),
        "descent_m": float(elev.descent_m),
        "max_gradient_pct": float(elev.max_gradient_pct),
        "avg_gradient_pct": float(elev.avg_gradient_pct),
        "max_above_start_m": float(elev.max_above_start_m),
        "max_below_start_m": float(elev.max_below_start_m),
        "end_diff_m": float(elev.end_diff_m),
        "green_percent": float(green.percent),
        "avg_trees_pct": float(green.avg_trees_pct),
        "avg_grass_pct": float(green.avg_grass_pct),
        "stress_cost_total": st["stress_cost_total"],
        "avg_stress_lts": st["avg_stress_lts"],
        "max_stress_lts": st["max_stress_lts"],
        "high_stress_segments_count": st["high_stress_segments_count"],
        "stressful_intersections_count": st["stressful_intersections_count"],
        "cost": float(r.cost),
        "mode": r.mode,
        "warnings_count": wn,
        "warnings_text": wt,
        "geometry_json": geom_json,
        "route_open_sky_share": round(float(getattr(r, "route_open_sky_share", 0.0)), 4),
        "route_building_shade_share": round(
            float(getattr(r, "route_building_shade_share", 0.0)), 4
        ),
        "route_covered_share": round(float(getattr(r, "route_covered_share", 0.0)), 4),
        "route_bad_wet_surface_share": round(
            float(getattr(r, "route_bad_wet_surface_share", 0.0)), 4
        ),
        "route_winter_harsh_surface_share": round(
            float(getattr(r, "route_winter_harsh_surface_share", 0.0)), 4
        ),
        "route_stairs_length_m": round(float(r.stairs.total_length_m), 2),
        "route_stairs_length_fraction": round(
            float(r.stairs.total_length_m) / max(float(r.length_m), 1.0), 4
        ),
        "route_winter_open_stress_proxy": round(
            float(getattr(r, "route_open_sky_share", 0.0))
            * float(
                getattr(getattr(r, "weather", None), "snow_model_strength", 0.0)
                or 0.0
            ),
            4,
        ),
        "route_wind_direction_aware": round(
            float(getattr(getattr(r, "heat_stress", None), "route_wind_direction_aware", 0.0)),
            4,
        ),
        "route_mean_wind_to_street_angle_deg": round(
            float(
                getattr(
                    getattr(r, "heat_stress", None),
                    "route_mean_wind_to_street_angle_deg",
                    0.0,
                )
            ),
            2,
        ),
        "route_mean_heat_directional_wind_exp": round(
            float(
                getattr(
                    getattr(r, "heat_stress", None),
                    "route_mean_heat_directional_wind_exp",
                    0.0,
                )
            ),
            4,
        ),
        "route_mean_heat_building_wind_factor": round(
            float(
                getattr(
                    getattr(r, "heat_stress", None),
                    "route_mean_heat_building_wind_factor",
                    0.0,
                )
            ),
            4,
        ),
        "route_frac_wind_along_open_hostile": round(
            float(
                getattr(
                    getattr(r, "heat_stress", None),
                    "route_frac_wind_along_open_hostile",
                    0.0,
                )
            ),
            4,
        ),
        "route_frac_wind_cross_building_screen": round(
            float(
                getattr(
                    getattr(r, "heat_stress", None),
                    "route_frac_wind_cross_building_screen",
                    0.0,
                )
            ),
            4,
        ),
        **{
            f"route_mean_{k}": round(
                float(
                    getattr(
                        getattr(r, "heat_stress", None),
                        f"route_mean_{k}",
                        0.0,
                    )
                ),
                4,
            )
            for k in CRITERION_KEYS
        },
        "stress_route_regime_factor": round(
            float(
                getattr(
                    getattr(r, "heat_stress", None),
                    "stress_route_regime_factor",
                    0.0,
                )
            ),
            4,
        ),
        "weather_test_case_id": None,
        "weather_test_temperature_c": None,
        "weather_test_precipitation_mm": None,
        "weather_test_wind_speed_ms": None,
        "weather_test_wind_gusts_ms": None,
        "weather_test_wind_direction_deg": None,
        "weather_test_cloud_cover_pct": None,
        "weather_test_humidity_pct": None,
        "weather_test_shortwave_radiation_wm2": None,
        "weather_test_snowfall_cm_h": None,
        "weather_test_snow_depth_m": None,
        "weather_test_weather_code": None,
        "weather_test_time_iso": None,
    }
    if test_weather_meta:
        out.update(test_weather_meta)
    if baseline_full is not None:
        out["full_baseline_length_m"] = round(float(baseline_full.length_m), 2)
        be = baseline_full.elevation
        out["delta_length_vs_full_m"] = round(
            float(r.length_m) - float(baseline_full.length_m), 2
        )
        out["delta_climb_vs_full_m"] = round(
            float(elev.climb_m) - float(be.climb_m), 2
        )
    else:
        out["full_baseline_length_m"] = None
        out["delta_length_vs_full_m"] = None
        out["delta_climb_vs_full_m"] = None
    return out


def _mean(xs: List[float]) -> Optional[float]:
    if not xs:
        return None
    return sum(xs) / len(xs)


def _write_sheet_table(
    ws: Any, headers: Sequence[str], rows: List[Dict[str, Any]]
) -> None:
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for ri, row in enumerate(rows, start=2):
        for c, h in enumerate(headers, start=1):
            v = row.get(h)
            ws.cell(row=ri, column=c, value=v)


def _make_summary_row(prefix: Dict[str, Any], rs: List[Dict[str, Any]]) -> Dict[str, Any]:
    row = dict(prefix)
    for k in SUMMARY_NUM_KEYS:
        vals: List[float] = []
        for r in rs:
            v = r.get(k)
            if v is None or v == "":
                continue
            try:
                vals.append(float(v))
            except (TypeError, ValueError):
                continue
        m = _mean(vals)
        row[f"mean_{k}"] = round(m, 4) if m is not None else None
    row["routes_count"] = len(rs)
    return row


def build_summaries(
    raw_rows: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """SummaryByVariant, SummaryByProfile, SummaryDirection."""

    def make_row(prefix: Dict[str, Any], rs: List[Dict[str, Any]]) -> Dict[str, Any]:
        return _make_summary_row(prefix, rs)

    by_pv: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    by_prof: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    by_dir: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for r in raw_rows:
        by_pv[(str(r["profile"]), str(r["variant_key"]))].append(r)
        by_prof[str(r["profile"])].append(r)
        by_dir[str(r["direction_order"])].append(r)

    s_var = [
        make_row({"profile": pk, "variant_key": vk}, rs)
        for (pk, vk), rs in sorted(by_pv.items())
    ]
    s_prof = [make_row({"profile": pk}, rs) for pk, rs in sorted(by_prof.items())]
    s_dir = [
        make_row({"direction_order": dk}, rs) for dk, rs in sorted(by_dir.items())
    ]
    return s_var, s_prof, s_dir


def build_summaries_by_weather_date(
    raw_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Средние по календарной дате погоды (multi-day)."""
    by_d: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        wd = str(r.get("weather_date") or "").strip()
        if wd:
            by_d[wd].append(r)
    return [
        _make_summary_row({"weather_date": d}, rs)
        for d, rs in sorted(by_d.items())
    ]


def build_summaries_by_weather_date_and_variant(
    raw_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Средние по паре (дата погоды, вариант маршрута)."""
    g: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        wd = str(r.get("weather_date") or "").strip()
        vk = str(r.get("variant_key") or "")
        if wd:
            g[(wd, vk)].append(r)
    return [
        _make_summary_row({"weather_date": a, "variant_key": b}, rs)
        for (a, b), rs in sorted(g.items())
    ]


def build_pair_comparison(
    raw_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Сводка по (origin, dest, profile[, дата погоды[, synthetic case]]): min/max/avg."""
    has_wd = bool(raw_rows and str(raw_rows[0].get("weather_date") or "").strip())
    has_tc = bool(raw_rows and str(raw_rows[0].get("weather_test_case_id") or "").strip())
    groups: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        base = (
            r["origin_point_id"],
            r["destination_point_id"],
            str(r["profile"]),
            str(r["direction_key"]),
        )
        el: List[Any] = list(base)
        if has_wd:
            el.append(str(r.get("weather_date") or "").strip())
        if has_tc:
            el.append(str(r.get("weather_test_case_id") or "").strip())
        key = tuple(el)
        groups[key].append(r)

    out: List[Dict[str, Any]] = []
    for key in sorted(groups.keys()):
        rs = groups[key]
        o, d, prof, dk = key[0], key[1], key[2], key[3]
        ii = 4
        wdate = str(key[ii]) if has_wd else ""
        if has_wd:
            ii += 1
        case_id = str(key[ii]) if has_tc else ""
        lengths = [float(x["length_m"]) for x in rs]
        times = [float(x["time_s"]) for x in rs]
        row: Dict[str, Any] = {
            "origin_point_id": o,
            "destination_point_id": d,
            "profile": prof,
            "direction_key": dk,
            "variants_count": len(rs),
            "length_min_m": min(lengths) if lengths else None,
            "length_max_m": max(lengths) if lengths else None,
            "length_mean_m": _mean(lengths),
            "time_min_s": min(times) if times else None,
            "time_max_s": max(times) if times else None,
            "time_mean_s": _mean(times),
        }
        if has_wd:
            row["weather_date"] = wdate
        if has_tc:
            row["weather_test_case_id"] = case_id
        out.append(row)
    return out


def mp_resolve_pool_workers(max_workers: int) -> int:
    """Число процессов пула: 0 → auto ``min(6, CPU−1)``, иначе ``max(1, N)``."""
    from multiprocessing import cpu_count

    c = int(cpu_count() or 4)
    if int(max_workers) <= 0:
        return max(1, min(6, c - 1))
    return max(1, int(max_workers))


def mp_split_weather_grid_chunks(grid: List[Any], wchunk: int) -> List[List[Any]]:
    w = max(1, int(wchunk))
    return [grid[k : k + w] for k in range(0, len(grid), w)]


def build_summaries_by_weather_case_id(
    raw_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Средние по ``weather_test_case_id`` (все варианты и профили вместе)."""
    by_c: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        cid = r.get("weather_test_case_id")
        if not cid:
            continue
        by_c[str(cid)].append(r)
    return [
        _make_summary_row({"weather_test_case_id": k}, rs)
        for k, rs in sorted(by_c.items())
    ]


def build_summaries_by_weather_case_and_variant(
    raw_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Средние по (synthetic-кейс, вариант маршрута)."""
    g: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        cid = r.get("weather_test_case_id")
        if not cid:
            continue
        vk = str(r.get("variant_key") or "")
        g[(str(cid), vk)].append(r)
    return [
        _make_summary_row({"weather_test_case_id": a, "variant_key": b}, rs)
        for (a, b), rs in sorted(g.items())
    ]


def _mean_finite(xs: List[float]) -> Optional[float]:
    vals = [x for x in xs if isinstance(x, (int, float)) and math.isfinite(float(x))]
    if not vals:
        return None
    return float(sum(vals)) / len(vals)


def build_heat_weather_influence_rows(
    raw_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """По каждому synthetic-кейсу: средние по heat и средние дельты heat−full по совпадающим O–D×профиль."""
    by_case: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        cid = r.get("weather_test_case_id")
        if not cid:
            continue
        by_case[str(cid)].append(r)

    out: List[Dict[str, Any]] = []
    for cid in sorted(by_case.keys()):
        rows = by_case[cid]
        trip_modes: Dict[Tuple[str, str, str], Dict[str, Dict[str, Any]]] = (
            defaultdict(dict)
        )
        for r in rows:
            o = str(r.get("origin_point_id") or "")
            d = str(r.get("destination_point_id") or "")
            p = str(r.get("profile") or "")
            v = str(r.get("variant_key") or "")
            trip_modes[(o, d, p)][v] = r

        heat_rows = [r for r in rows if str(r.get("variant_key")) == "heat"]
        mean_len_h = _mean_finite([float(r["length_m"]) for r in heat_rows])
        mean_time_h = _mean_finite([float(r["time_s"]) for r in heat_rows])
        mean_climb_h = _mean_finite([float(r["climb_m"]) for r in heat_rows])
        mean_green_h = _mean_finite([float(r["green_percent"]) for r in heat_rows])
        mean_trees_h = _mean_finite([float(r["avg_trees_pct"]) for r in heat_rows])
        mean_stress_h = _mean_finite(
            [float(r["stress_cost_total"]) for r in heat_rows]
        )

        dlen: List[float] = []
        dtime: List[float] = []
        dgreen: List[float] = []
        dtrees: List[float] = []
        dstress: List[float] = []
        for _k, modes in trip_modes.items():
            h = modes.get("heat")
            f = modes.get("full")
            if not h or not f:
                continue
            dlen.append(float(h["length_m"]) - float(f["length_m"]))
            dtime.append(float(h["time_s"]) - float(f["time_s"]))
            dgreen.append(float(h["green_percent"]) - float(f["green_percent"]))
            dtrees.append(float(h["avg_trees_pct"]) - float(f["avg_trees_pct"]))
            dstress.append(float(h["stress_cost_total"]) - float(f["stress_cost_total"]))

        sample = heat_rows[0] if heat_rows else rows[0]
        row: Dict[str, Any] = {
            "weather_test_case_id": cid,
            "weather_test_temperature_c": sample.get("weather_test_temperature_c"),
            "weather_test_precipitation_mm": sample.get("weather_test_precipitation_mm"),
            "weather_test_wind_speed_ms": sample.get("weather_test_wind_speed_ms"),
            "weather_test_cloud_cover_pct": sample.get("weather_test_cloud_cover_pct"),
            "mean_length_heat_m": round(mean_len_h, 4) if mean_len_h is not None else None,
            "mean_time_heat_s": round(mean_time_h, 4) if mean_time_h is not None else None,
            "mean_climb_heat_m": round(mean_climb_h, 4) if mean_climb_h is not None else None,
            "mean_green_heat_pct": round(mean_green_h, 4) if mean_green_h is not None else None,
            "mean_trees_heat_pct": round(mean_trees_h, 4) if mean_trees_h is not None else None,
            "mean_stress_cost_heat": round(mean_stress_h, 4) if mean_stress_h is not None else None,
            "mean_delta_length_heat_minus_full_m": round(_mean_finite(dlen), 4)
            if dlen
            else None,
            "mean_delta_time_heat_minus_full_s": round(_mean_finite(dtime), 4)
            if dtime
            else None,
            "mean_delta_green_heat_minus_full_pct": round(_mean_finite(dgreen), 4)
            if dgreen
            else None,
            "mean_delta_trees_heat_minus_full_pct": round(_mean_finite(dtrees), 4)
            if dtrees
            else None,
            "mean_delta_stress_heat_minus_full": round(_mean_finite(dstress), 4)
            if dstress
            else None,
            "paired_od_profile_count": len(dlen),
        }
        out.append(row)
    return out


def build_heat_vs_green_by_weather_rows(
    raw_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """По synthetic-кейсу: средние дельты heat−green (длина, время, стресс, зелень)."""
    by_case: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        cid = r.get("weather_test_case_id")
        if not cid:
            continue
        by_case[str(cid)].append(r)

    out: List[Dict[str, Any]] = []
    for cid in sorted(by_case.keys()):
        rows = by_case[cid]
        trip_modes: Dict[Tuple[str, str, str], Dict[str, Dict[str, Any]]] = (
            defaultdict(dict)
        )
        for r in rows:
            o = str(r.get("origin_point_id") or "")
            d = str(r.get("destination_point_id") or "")
            p = str(r.get("profile") or "")
            v = str(r.get("variant_key") or "")
            trip_modes[(o, d, p)][v] = r

        dlen: List[float] = []
        dtime: List[float] = []
        dgreen: List[float] = []
        dstress: List[float] = []
        for _k, modes in trip_modes.items():
            h = modes.get("heat")
            g = modes.get("green")
            if not h or not g:
                continue
            dlen.append(float(h["length_m"]) - float(g["length_m"]))
            dtime.append(float(h["time_s"]) - float(g["time_s"]))
            dgreen.append(float(h["green_percent"]) - float(g["green_percent"]))
            dstress.append(float(h["stress_cost_total"]) - float(g["stress_cost_total"]))

        sample = next((r for r in rows if str(r.get("variant_key")) == "heat"), rows[0])
        out.append(
            {
                "weather_test_case_id": cid,
                "weather_test_temperature_c": sample.get("weather_test_temperature_c"),
                "mean_delta_length_heat_minus_green_m": round(_mean_finite(dlen), 4)
                if dlen
                else None,
                "mean_delta_time_heat_minus_green_s": round(_mean_finite(dtime), 4)
                if dtime
                else None,
                "mean_delta_green_heat_minus_green_pct": round(_mean_finite(dgreen), 4)
                if dgreen
                else None,
                "mean_delta_stress_heat_minus_green": round(_mean_finite(dstress), 4)
                if dstress
                else None,
                "paired_od_profile_count": len(dlen),
            }
        )
    return out


def _heat_row_float(r: Dict[str, Any], key: str) -> float:
    v = r.get(key)
    if v is None or v == "":
        return float("nan")
    try:
        return float(v)
    except (TypeError, ValueError):
        return float("nan")


def _heat_row_rain_flag(r: Dict[str, Any]) -> int:
    pr = _heat_row_float(r, "weather_test_precipitation_mm")
    if not math.isfinite(pr):
        return 0
    return 1 if pr > 0.5 else 0


def _heat_row_wind_strong(r: Dict[str, Any]) -> int:
    ws = _heat_row_float(r, "weather_test_wind_speed_ms")
    if not math.isfinite(ws):
        return 0
    return 1 if ws >= 6.0 else 0


def build_heat_weather_kpi_rows(raw_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Сводные KPI для heat_weather (36 synthetic): геометрия vs погода."""
    if not raw_rows:
        return [
            {
                "changed_pairs_count": 0,
                "changed_pairs_share": None,
                "mean_unique_geometries": None,
                "scenario_alt_rate_min": None,
                "scenario_alt_rate_max": None,
                "temp25_vs_0_open_lower_share": None,
                "rain_vs_dry_open_lower_share": None,
                "wind_vs_calm_open_lower_share": None,
                "covered_nonzero_share": None,
            }
        ]

    def od_prof(r: Dict[str, Any]) -> Tuple[str, str, str]:
        return (
            str(r["origin_point_id"]),
            str(r["destination_point_id"]),
            str(r["profile"]),
        )

    by_pair: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        cid = r.get("weather_test_case_id")
        if not cid:
            continue
        by_pair[od_prof(r)].append(r)

    n_groups = len(by_pair)
    unique_counts: List[int] = []
    changed = 0
    for _k, lst in by_pair.items():
        geoms = {str(x.get("geometry_json") or "") for x in lst}
        u = len(geoms)
        unique_counts.append(u)
        if u > 1:
            changed += 1
    mean_unique = (
        sum(unique_counts) / len(unique_counts) if unique_counts else None
    )
    changed_share = (changed / n_groups) if n_groups else None

    geom_ctr: Dict[Tuple[str, str, str], Counter] = defaultdict(Counter)
    for r in raw_rows:
        if not r.get("weather_test_case_id"):
            continue
        g = str(r.get("geometry_json") or "")
        geom_ctr[od_prof(r)][g] += 1
    modal: Dict[Tuple[str, str, str], str] = {}
    for k, ctr in geom_ctr.items():
        modal[k] = ctr.most_common(1)[0][0] if ctr else ""

    by_case: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        cid = r.get("weather_test_case_id")
        if cid:
            by_case[str(cid)].append(r)

    scenario_rates: List[float] = []
    for cid, lst in by_case.items():
        alt = 0
        tot = 0
        for r in lst:
            k = od_prof(r)
            tot += 1
            g = str(r.get("geometry_json") or "")
            if g != modal.get(k, ""):
                alt += 1
        if tot:
            scenario_rates.append(alt / tot)
    scen_min = min(scenario_rates) if scenario_rates else None
    scen_max = max(scenario_rates) if scenario_rates else None

    bucket_t: Dict[Tuple[int, int, int, str, str, str], Dict[float, Dict[str, Any]]] = (
        defaultdict(dict)
    )
    bucket_r: Dict[Tuple[float, int, int, str, str, str], Dict[int, Dict[str, Any]]] = (
        defaultdict(dict)
    )
    bucket_w: Dict[Tuple[float, int, int, str, str, str], Dict[int, Dict[str, Any]]] = (
        defaultdict(dict)
    )

    for r in raw_rows:
        if not r.get("weather_test_case_id"):
            continue
        o, d, p = od_prof(r)
        T = round(_heat_row_float(r, "weather_test_temperature_c"), 1)
        cc = int(round(_heat_row_float(r, "weather_test_cloud_cover_pct")))
        rf = _heat_row_rain_flag(r)
        wf = _heat_row_wind_strong(r)
        open_v = _heat_row_float(r, "route_open_sky_share")

        kt = (rf, wf, cc, o, d, p)
        if T in (0.0, 25.0) and math.isfinite(open_v):
            dct = bucket_t[kt]
            dct[T] = r

        kr = (T, wf, cc, o, d, p)
        if rf in (0, 1) and math.isfinite(open_v):
            dcr = bucket_r[kr]
            dcr[rf] = r

        kw = (T, rf, cc, o, d, p)
        if wf in (0, 1) and math.isfinite(open_v):
            dcw = bucket_w[kw]
            dcw[wf] = r

    def _pair_open_lower(
        d: Dict[Any, Dict[Any, Dict[str, Any]]], hi: Any, lo: Any
    ) -> Optional[float]:
        ok = 0
        good = 0
        for _bk, dct in d.items():
            if hi not in dct or lo not in dct:
                continue
            rh, rl = dct[hi], dct[lo]
            oh = _heat_row_float(rh, "route_open_sky_share")
            ol = _heat_row_float(rl, "route_open_sky_share")
            if not (math.isfinite(oh) and math.isfinite(ol)):
                continue
            ok += 1
            if oh < ol - 1e-9:
                good += 1
        return (good / ok) if ok else None

    temp_share = _pair_open_lower(bucket_t, 25.0, 0.0)
    rain_share = _pair_open_lower(bucket_r, 1, 0)
    wind_share = _pair_open_lower(bucket_w, 1, 0)

    cov_ok = sum(
        1
        for r in raw_rows
        if _heat_row_float(r, "route_covered_share") > 1e-6
    )
    covered_share = cov_ok / len(raw_rows) if raw_rows else None

    row = {
        "changed_pairs_count": changed,
        "changed_pairs_share": round(changed_share, 6) if changed_share is not None else None,
        "mean_unique_geometries": round(mean_unique, 6) if mean_unique is not None else None,
        "scenario_alt_rate_min": round(scen_min, 6) if scen_min is not None else None,
        "scenario_alt_rate_max": round(scen_max, 6) if scen_max is not None else None,
        "temp25_vs_0_open_lower_share": round(temp_share, 6)
        if temp_share is not None
        else None,
        "rain_vs_dry_open_lower_share": round(rain_share, 6)
        if rain_share is not None
        else None,
        "wind_vs_calm_open_lower_share": round(wind_share, 6)
        if wind_share is not None
        else None,
        "covered_nonzero_share": round(covered_share, 6)
        if covered_share is not None
        else None,
    }
    return [row]


def _heat_od_prof(r: Dict[str, Any]) -> Tuple[str, str, str]:
    return (
        str(r.get("origin_point_id") or ""),
        str(r.get("destination_point_id") or ""),
        str(r.get("profile") or ""),
    )


def _weather_case_wind_stem(case_id: str) -> str:
    """Базовый id сценария без суффикса ``_WD{deg}`` (лето/зима с направлением ветра)."""
    s = str(case_id)
    m = re.search(r"_WD\d+$", s)
    return s[: m.start()] if m else s


def _col_finite_floats(rows: List[Dict[str, Any]], key: str) -> List[float]:
    out: List[float] = []
    for r in rows:
        x = _heat_row_float(r, key)
        if math.isfinite(x):
            out.append(x)
    return out


def build_heat_weather_kpi_extras_dict(raw_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Поля для merge в первую строку Heat-weather KPI: зелёный критерий, ветер, покрытие, surface."""
    ex: Dict[str, Any] = {}
    if not raw_rows:
        return ex

    gr = _col_finite_floats(raw_rows, "route_mean_green_route_factor")
    if gr:
        ex["green_route_factor_min"] = round(min(gr), 6)
        ex["green_route_factor_max"] = round(max(gr), 6)
        ex["green_route_factor_mean"] = round(sum(gr) / len(gr), 6)
        mu = sum(gr) / len(gr)
        ex["green_route_factor_std"] = round(
            math.sqrt(sum((x - mu) ** 2 for x in gr) / len(gr)), 6
        ) if len(gr) > 1 else 0.0
    else:
        ex["green_route_factor_min"] = None
        ex["green_route_factor_max"] = None
        ex["green_route_factor_mean"] = None
        ex["green_route_factor_std"] = None

    ex["green_route_factor_heat_note_ru"] = (
        "При включённой непрерывной тепло-модели (heat_continuous) тень деревьев и "
        "микроклимат в основном входят в ef heat (open/tree/build), а столбец "
        "route_mean_green_route_factor — это только дискретный бонус политики "
        "max(wm.green−1,0)×trees_pct×green_coupling на ребре; при wm.green≤1 или "
        "низкой доле деревьев на пути он остаётся ≈1.0 без ошибки расчёта."
    )

    by_stem_od: Dict[Tuple[str, str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        cid = r.get("weather_test_case_id")
        if not cid:
            continue
        stem = _weather_case_wind_stem(str(cid))
        if stem == str(cid):
            continue
        by_stem_od[(stem, *_heat_od_prof(r))].append(r)

    wind_geom_groups = 0
    for _k, lst in by_stem_od.items():
        if len(lst) < 2:
            continue
        geoms = {str(x.get("geometry_json") or "") for x in lst}
        if len(geoms) > 1:
            wind_geom_groups += 1
    ex["wind_dir_only_geometry_change_groups"] = int(wind_geom_groups)

    orient_ranges: List[float] = []
    orient_notable_pairs = 0
    n_pairs_measured = 0
    by_od: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        if not r.get("weather_test_case_id"):
            continue
        by_od[_heat_od_prof(r)].append(r)
    for _od, lst in by_od.items():
        wof = _col_finite_floats(lst, "route_mean_wind_orientation_factor")
        if len(wof) < 2:
            continue
        n_pairs_measured += 1
        rng = max(wof) - min(wof)
        orient_ranges.append(rng)
        if rng >= 0.02:
            orient_notable_pairs += 1
    ex["mean_wind_orientation_factor_range_per_od_pair"] = (
        round(sum(orient_ranges) / len(orient_ranges), 6) if orient_ranges else None
    )
    ex["share_od_pairs_wind_orientation_range_ge_0p02"] = (
        round(orient_notable_pairs / n_pairs_measured, 6) if n_pairs_measured else None
    )

    stairs_spreads: List[float] = []
    harsh_spreads: List[float] = []
    shade_spreads: List[float] = []
    for _od, lst in by_od.items():
        if len(lst) < 2:
            continue
        for key, bucket in (
            ("route_stairs_length_fraction", stairs_spreads),
            ("route_winter_harsh_surface_share", harsh_spreads),
            ("route_building_shade_share", shade_spreads),
        ):
            vs = _col_finite_floats(lst, key)
            if len(vs) >= 2:
                bucket.append(max(vs) - min(vs))
    ex["coverage_max_stairs_fraction_spread_any_od_pair"] = (
        round(max(stairs_spreads), 6) if stairs_spreads else None
    )
    ex["coverage_max_winter_harsh_surface_spread_any_od_pair"] = (
        round(max(harsh_spreads), 6) if harsh_spreads else None
    )
    ex["coverage_max_building_shade_spread_any_od_pair"] = (
        round(max(shade_spreads), 6) if shade_spreads else None
    )
    ex["coverage_any_row_covered_share_positive"] = int(
        any(_heat_row_float(r, "route_covered_share") > 1e-5 for r in raw_rows)
    )
    ex["coverage_any_pair_stairs_spread_ge_0p002"] = int(
        bool(stairs_spreads and max(stairs_spreads) >= 0.002)
    )
    ex["coverage_any_pair_harsh_spread_ge_0p02"] = int(
        bool(harsh_spreads and max(harsh_spreads) >= 0.02)
    )
    ex["coverage_any_pair_building_shade_spread_ge_0p02"] = int(
        bool(shade_spreads and max(shade_spreads) >= 0.02)
    )

    surf = _col_finite_floats(raw_rows, "route_mean_surface_weather_factor")
    wet = _col_finite_floats(raw_rows, "route_bad_wet_surface_share")
    if surf:
        ex["surface_weather_factor_min"] = round(min(surf), 6)
        ex["surface_weather_factor_max"] = round(max(surf), 6)
        ex["surface_weather_factor_mean"] = round(sum(surf) / len(surf), 6)
    else:
        ex["surface_weather_factor_min"] = None
        ex["surface_weather_factor_max"] = None
        ex["surface_weather_factor_mean"] = None
    if wet:
        ex["bad_wet_surface_share_min"] = round(min(wet), 6)
        ex["bad_wet_surface_share_max"] = round(max(wet), 6)
        ex["bad_wet_surface_share_mean"] = round(sum(wet) / len(wet), 6)
    else:
        ex["bad_wet_surface_share_min"] = None
        ex["bad_wet_surface_share_max"] = None
        ex["bad_wet_surface_share_mean"] = None

    parts: List[str] = []
    if ex.get("coverage_any_row_covered_share_positive"):
        parts.append("есть ненулевой covered_share")
    if ex.get("coverage_any_pair_stairs_spread_ge_0p002"):
        parts.append("есть разброс доли лестниц по парам")
    if ex.get("coverage_any_pair_harsh_spread_ge_0p02"):
        parts.append("есть разброс winter_harsh_surface по парам")
    if ex.get("coverage_any_pair_building_shade_spread_ge_0p02"):
        parts.append("есть разброс building_shade по парам")
    ex["validation_coverage_summary_ru"] = (
        "; ".join(parts) if parts else "слабое покрытие: мало вариации stairs/harsh/shade/covered"
    )

    return ex


def build_heat_qa_warning_rows(raw_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Автоматические QA-предупреждения по константности и «невидимому» reroute."""
    warns: List[Dict[str, Any]] = []
    if not raw_rows:
        warns.append({"severity": "info", "code": "empty", "message": "нет строк маршрутов"})
        return warns

    gr = _col_finite_floats(raw_rows, "route_mean_green_route_factor")
    if gr and max(gr) - min(gr) < 1e-4:
        warns.append(
            {
                "severity": "info",
                "code": "green_route_factor_constant",
                "message": "route_mean_green_route_factor почти не меняется по строкам; "
                "для heat см. green_route_factor_heat_note_ru в KPI (не баг сезонного множителя).",
            }
        )

    cov = _col_finite_floats(raw_rows, "route_covered_share")
    if cov and max(cov) < 1e-6:
        warns.append(
            {
                "severity": "warning",
                "code": "covered_share_all_zero",
                "message": "route_covered_share везде ~0 — выборка не валидирует навесы/covered.",
            }
        )

    bs = _col_finite_floats(raw_rows, "route_building_shade_share")
    if len(bs) > 3 and max(bs) - min(bs) < 0.008:
        warns.append(
            {
                "severity": "info",
                "code": "building_shade_low_spread",
                "message": "Низкий разброс route_building_shade_share по всему батчу.",
            }
        )

    sf = _col_finite_floats(raw_rows, "route_stairs_length_fraction")
    if len(sf) > 3 and max(sf) - min(sf) < 0.001:
        warns.append(
            {
                "severity": "info",
                "code": "stairs_fraction_low_spread",
                "message": "Доля лестниц почти не варьируется — мало чувствительности к лестницам.",
            }
        )

    wh = _col_finite_floats(raw_rows, "route_winter_harsh_surface_share")
    if len(wh) > 5 and max(wh) - min(wh) < 0.015:
        warns.append(
            {
                "severity": "info",
                "code": "winter_harsh_low_spread",
                "message": "Низкий разброс route_winter_harsh_surface_share (или мало зимних строк).",
            }
        )

    by_od: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        if r.get("weather_test_case_id"):
            by_od[_heat_od_prof(r)].append(r)
    invisible_reroute_pairs = 0
    multi_geom_pairs = 0
    for lst in by_od.values():
        if len(lst) < 2:
            continue
        geoms = {str(x.get("geometry_json") or "") for x in lst}
        if len(geoms) <= 1:
            continue
        multi_geom_pairs += 1
        metric_keys = (
            "route_stairs_length_fraction",
            "route_winter_harsh_surface_share",
            "route_open_sky_share",
            "route_building_shade_share",
            "route_mean_stress_weather_factor",
        )
        max_spread = 0.0
        for mk in metric_keys:
            vs = _col_finite_floats(lst, mk)
            if len(vs) >= 2:
                max_spread = max(max_spread, max(vs) - min(vs))
        if max_spread < 0.02:
            invisible_reroute_pairs += 1
    if invisible_reroute_pairs:
        warns.append(
            {
                "severity": "warning",
                "code": "reroute_low_visible_metric_delta",
                "message": (
                    f"У {invisible_reroute_pairs} O-D-профилей с >1 геометрией макс. разброс "
                    "stairs/harsh/open/building/stress < 0.02 — reroute плохо читается по этим полям."
                ),
            }
        )
    if multi_geom_pairs == 0:
        warns.append(
            {
                "severity": "info",
                "code": "no_geometry_switch",
                "message": "Ни одна O-D-пара не меняет geometry_json между сценариями.",
            }
        )

    if not warns:
        warns.append({"severity": "ok", "code": "none", "message": "Явных QA-флагов нет."})
    return warns


def build_winter_reroute_explain_rows(raw_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """По зимним строкам: для пар с >1 геометрией — диапазоны метрик и сильнейший относительный сигнал."""
    wrows = [
        r
        for r in raw_rows
        if str(r.get("weather_test_case_id") or "").startswith("winter_")
    ]
    if not wrows:
        return [
            {
                "note": "нет зимних synthetic-строк (weather_test_case_id не с префикса winter_)",
            }
        ]

    by_pair: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for r in wrows:
        by_pair[_heat_od_prof(r)].append(r)

    metrics = (
        ("stairs_fraction", "route_stairs_length_fraction"),
        ("winter_harsh_surface", "route_winter_harsh_surface_share"),
        ("open_sky", "route_open_sky_share"),
        ("building_shade", "route_building_shade_share"),
        ("stress_weather", "route_mean_stress_weather_factor"),
    )
    out: List[Dict[str, Any]] = []
    for key, lst in by_pair.items():
        geoms = {str(x.get("geometry_json") or "") for x in lst if x.get("geometry_json")}
        if len(geoms) < 2:
            continue
        row: Dict[str, Any] = {
            "origin_point_id": key[0],
            "destination_point_id": key[1],
            "profile": key[2],
            "winter_distinct_geometries": len(geoms),
        }
        best_label = ""
        best_rel = -1.0
        for label, mk in metrics:
            vs = _col_finite_floats(lst, mk)
            if len(vs) < 2:
                rng = 0.0
                mabs = 0.0
            else:
                rng = max(vs) - min(vs)
                mabs = sum(abs(v) for v in vs) / len(vs)
            row[f"range_{label}"] = round(rng, 6)
            rel = rng / max(0.05, mabs) if mabs > 1e-9 else rng
            if rel > best_rel:
                best_rel = rel
                best_label = label
        row["strongest_relative_range_metric"] = best_label or "n/a"
        out.append(row)

    if not out:
        return [
            {
                "note": "есть зимние строки, но ни одна O-D-пара не дала >1 геометрии",
            }
        ]
    return out


def build_heat_experiment_extra_sheet_blocks(
    raw_rows: List[Dict[str, Any]],
) -> List[Tuple[str, List[Dict[str, Any]]]]:
    """Доп. блоки для листа «Сводка» (только heat-эксперимент)."""
    return [
        ("QA — предупреждения", build_heat_qa_warning_rows(raw_rows)),
        ("Зима: объяснение reroute (диапазоны метрик)", build_winter_reroute_explain_rows(raw_rows)),
    ]


def build_winter_kpi_rows(raw_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Сводные KPI по зимним synthetic-строкам (``weather_test_case_id``)."""
    wr = [r for r in raw_rows if str(r.get("weather_test_case_id") or "").strip()]
    if not wr:
        return [
            {
                "winter_kpi_note": "нет строк с weather_test_case_id",
                "winter_synthetic_routes": 0,
            }
        ]

    def _avg(rows: List[Dict[str, Any]], key: str) -> Optional[float]:
        xs: List[float] = []
        for r in rows:
            v = r.get(key)
            if v is None or v == "":
                continue
            try:
                xf = float(v)
            except (TypeError, ValueError):
                continue
            if math.isfinite(xf):
                xs.append(xf)
        return sum(xs) / len(xs) if xs else None

    high_snow = [
        r
        for r in wr
        if _heat_row_float(r, "weather_test_snow_depth_m") >= 0.05
        or _heat_row_float(r, "weather_test_snowfall_cm_h") >= 0.35
    ]
    wind_snow = [
        r
        for r in wr
        if _heat_row_float(r, "weather_test_wind_speed_ms") >= 8.0
        and (
            _heat_row_float(r, "weather_test_snowfall_cm_h") > 0.15
            or _heat_row_float(r, "weather_test_snow_depth_m") >= 0.03
        )
    ]

    def _cpkm(r: Dict[str, Any]) -> float:
        return float(r.get("cost") or 0.0) / max(float(r.get("length_km") or 1e-9), 1e-9)

    c_list = [_cpkm(r) for r in high_snow if str(r.get("profile")) == "cyclist"]
    p_list = [_cpkm(r) for r in high_snow if str(r.get("profile")) == "pedestrian"]
    c_avg = sum(c_list) / len(c_list) if c_list else None
    p_avg = sum(p_list) / len(p_list) if p_list else None
    ratio = (c_avg / p_avg) if (c_avg is not None and p_avg and p_avg > 1e-9) else None

    april_weak = sum(
        1
        for r in wr
        if "2000-04-05" in str(r.get("weather_test_time_iso") or "")
        and float(r.get("weather_season_green_mult") or 1.0) < 0.35
    )
    april_strong = sum(
        1
        for r in wr
        if "2000-04-25" in str(r.get("weather_test_time_iso") or "")
        and float(r.get("weather_season_green_mult") or 0.0) > 0.85
    )
    april_rows_early = [
        r
        for r in wr
        if "2000-04-05" in str(r.get("weather_test_time_iso") or "")
    ]
    april_rows_late = [
        r
        for r in wr
        if "2000-04-25" in str(r.get("weather_test_time_iso") or "")
    ]
    mixed_slip_wc: List[Dict[str, Any]] = []
    for r in wr:
        wcv = r.get("weather_test_weather_code")
        if wcv is None or wcv == "":
            continue
        try:
            ci = int(round(float(wcv)))
        except (TypeError, ValueError):
            continue
        if ci in (56, 57, 66, 67):
            mixed_slip_wc.append(r)

    return [
        {
            "winter_synthetic_routes": len(wr),
            "mean_stairs_len_frac_high_snow": round(
                _avg(high_snow, "route_stairs_length_fraction") or 0.0, 6
            )
            if high_snow
            else None,
            "mean_harsh_surface_share_high_snow": round(
                _avg(high_snow, "route_winter_harsh_surface_share") or 0.0, 6
            )
            if high_snow
            else None,
            "mean_open_sky_share_wind_and_snow": round(
                _avg(wind_snow, "route_open_sky_share") or 0.0, 6
            )
            if wind_snow
            else None,
            "mean_building_shade_share_all_winter_rows": round(
                _avg(wr, "route_building_shade_share") or 0.0, 6
            ),
            "april_early_green_mult_weak_count": april_weak,
            "april_late_green_mult_strong_count": april_strong,
            "cyclist_vs_ped_cost_per_km_ratio_high_snow": round(ratio, 6)
            if ratio is not None
            else None,
            "mean_route_mean_surface_weather_factor_high_snow": round(
                _avg(high_snow, "route_mean_surface_weather_factor") or 0.0, 6
            )
            if high_snow
            else None,
            "mean_route_mean_stairs_weather_factor_high_snow": round(
                _avg(high_snow, "route_mean_stairs_weather_factor") or 0.0, 6
            )
            if high_snow
            else None,
            "mean_route_mean_green_route_factor_april_early": round(
                _avg(april_rows_early, "route_mean_green_route_factor") or 0.0, 6
            )
            if april_rows_early
            else None,
            "mean_route_mean_green_route_factor_april_late": round(
                _avg(april_rows_late, "route_mean_green_route_factor") or 0.0, 6
            )
            if april_rows_late
            else None,
            "mean_route_mean_wind_orientation_factor_wind_snow": round(
                _avg(wind_snow, "route_mean_wind_orientation_factor") or 0.0, 6
            )
            if wind_snow
            else None,
            "mean_route_mean_stress_weather_factor_mixed_wc_rows": round(
                _avg(mixed_slip_wc, "route_mean_stress_weather_factor") or 0.0, 6
            )
            if mixed_slip_wc
            else None,
            "mean_stress_route_regime_factor_all_winter_rows": round(
                _avg(wr, "stress_route_regime_factor") or 0.0, 6
            ),
        }
    ]


def _meta_rows_ru(rows: List[Tuple[str, Any]]) -> List[Tuple[str, Any]]:
    ru = {
        "experiment_id": "ID эксперимента",
        "started_at_utc": "Время старта (UTC)",
        "seed": "Seed",
        "n_points": "Число точек",
        "n_directed_pairs": "Направленных пар O–D",
        "n_od_pairs": "Число пар O–D в батче (направленных или неориентированных)",
        "heat_route_tasks": "Число задач heat (пара×профиль)",
        "heat_directed_pairs": "Heat: directed-пары A→B и B→A",
        "heat_max_workers": "Heat: число процессов пула (1=последовательно)",
        "heat_mp_weather_chunk": "Heat MP: synthetic-кейсов в одной задаче пула",
        "batch_max_workers": "Батч: процессов пула (1=последовательно)",
        "batch_chunk_size": "Батч: chunksize для pool.imap",
        "batch_mp_weather_chunk": "Батч: synthetic-кейсов на задачу пула",
        "batch_mp_mode": "Батч: режим ускорения (parallel / sequential)",
        "batch_directed_pairs": "Батч: направленные пары A→B и B→A",
        "n_weather_cases": "Число synthetic-кейсов погоды",
        "expected_variant_rows": "Ожидалось строк маршрутов (кейсы×пары×профили×6)",
        "batch_profiles": "Профили (батч)",
        "n_profiles": "Профилей",
        "n_variants": "Вариантов маршрута",
        "expected_route_cells": "Ожидалось строк маршрутов",
        "successful_route_rows": "Успешных строк",
        "skipped_od_no_path_buffer_100m": "Пропущено (нет пути до 100 м)",
        "failure_rows": "Строк ошибок",
        "precache_polygon_wkt_sha256": "SHA256 полигона precache",
        "precache_polygon_bounds_lonlat": "Границы полигона (lon/lat)",
        "precache_polygon_wkt_prefix": "Префикс WKT полигона",
        "routing_algo_version": "Версия алгоритма",
        "routing_weights_fingerprint": "Отпечаток весов",
        "elapsed_seconds": "Время работы, с",
        "weather_schedule": "Сценарий погоды (none / now / past / test)",
        "synthetic_test_cases": "Число синтетических сценариев (test)",
        "experiment_weather_grid": "Сетка synthetic (summer / winter / all)",
        "weather_mode_engine": "Режим погоды (движок)",
        "weather_time_utc_snapshot": "Снимок времени UTC (режим now)",
        "past_days": "Дней архива (режим past)",
        "past_slot_local_hour": "Локальный час слота (режим past)",
        "batch_weather_tz": "Часовой пояс IANA (режим past)",
        "vertices_file": "Файл вершин (отдельно)",
        "experiment_kind": "Тип эксперимента",
        "heat_continuous_enable": "Непрерывная heat-модель на ребре (HEAT_CONTINUOUS_ENABLE)",
        "batch_weather_center_lat": "Центр запроса погоды: широта",
        "batch_weather_center_lon": "Центр запроса погоды: долгота",
        "batch_weather_departure_iso": "ISO времени для погоды (батч)",
        "snapshot_temperature_c": "Снимок: температура, °C",
        "snapshot_apparent_temperature_c": "Снимок: ощущается, °C",
        "snapshot_precipitation_mm": "Снимок: осадки, мм/ч",
        "snapshot_precipitation_probability": "Снимок: вероятность осадков, %",
        "snapshot_wind_speed_ms": "Снимок: ветер, м/с",
        "snapshot_wind_gusts_ms": "Снимок: порывы, м/с",
        "snapshot_wind_direction_deg": "Снимок: направление ветра, ° (откуда дует)",
        "snapshot_cloud_cover_pct": "Снимок: облачность, %",
        "snapshot_humidity_pct": "Снимок: влажность, %",
        "snapshot_shortwave_radiation_wm2": "Снимок: КВ, Вт/м²",
        "snapshot_snowfall_cm_h": "Снимок: снегопад, см/ч",
        "snapshot_snow_depth_m": "Снимок: глубина снега, м",
        "snapshot_weather_code": "Снимок: код погоды WMO",
        "snapshot_routing_season": "Снимок: эффективный сезон маршрутизации",
        "snapshot_routing_season_calendar": "Снимок: календарный сезон",
        "snapshot_routing_season_source": "Снимок: источник сезона (calendar/adaptive)",
        "snapshot_season_green_mult": "Снимок: множитель зелёного бонуса",
        "snapshot_snow_model_strength": "Снимок: сила зимней snow-модели",
        "weather_snapshot_source": "Источник погодного снимка (батч)",
        "synthetic_weather_time_iso": "Synthetic: ISO времени снимка",
        "project_weather_stress_global_blend": "Проект: WEATHER_STRESS_GLOBAL_BLEND",
        "experiment_weather_stress_global_blend_effective": "Эксперимент: blend (эффективно)",
        "experiment_stress_blend_mode": "Режим blend (метка)",
    }
    return [(ru.get(str(k), str(k)), v) for k, v in rows]


_ROUTE_COL_RU: Dict[str, str] = {
    "route_id": "ID маршрута",
    "experiment_id": "ID эксперимента",
    "seed": "Seed",
    "profile": "Профиль",
    "variant_key": "Вариант (ключ)",
    "variant_label": "Вариант",
    "origin_point_id": "Точка отправления",
    "destination_point_id": "Точка назначения",
    "direction_key": "Направление",
    "direction_order": "Порядок",
    "origin_lat": "Широта начала",
    "origin_lon": "Долгота начала",
    "destination_lat": "Широта конца",
    "destination_lon": "Долгота конца",
    "route_built_at_utc": "Построен в (UTC)",
    "weather_date": "Дата погоды",
    "weather_time": "Время погоды (ISO)",
    "weather_source": "Источник погоды",
    "weather_temperature_c": "Температура, °C",
    "weather_apparent_temperature_c": "Ощущается, °C",
    "weather_precipitation_mm": "Осадки, мм/ч",
    "weather_precipitation_probability": "Вероятность осадков, %",
    "weather_wind_speed_ms": "Ветер, м/с",
    "weather_wind_gusts_ms": "Порывы ветра, м/с",
    "weather_wind_direction_deg": "Направление ветра, ° (метео: откуда)",
    "weather_wind_direction_aware": "Ветер с направлением (direction-aware)",
    "weather_cloud_cover_pct": "Облачность, %",
    "weather_humidity_pct": "Влажность, %",
    "weather_shortwave_radiation_wm2": "КВ радиация, Вт/м²",
    "weather_snowfall_cm_h": "Снегопад, см/ч",
    "weather_snow_depth_m": "Глубина снега, м (модель)",
    "weather_weather_code": "Код погоды WMO",
    "weather_routing_season": "Сезонный профиль маршрутизации (эффективный)",
    "weather_routing_season_calendar": "Сезон по календарю",
    "weather_routing_season_source": "Источник сезона (calendar/adaptive)",
    "weather_season_green_mult": "Сезон: множитель зелёного бонуса",
    "weather_season_tree_heat_mult": "Сезон: множитель тени деревьев (heat)",
    "weather_season_stress_route_mult": "Сезон: множитель глобального stress",
    "weather_season_stairs_route_mult": "Сезон: множитель критерия лестниц",
    "weather_season_wind_orientation_route_mult": "Сезон: множитель ориентации к ветру",
    "weather_stress_route_regime_factor": "Глобальный stress-regime (blend+snow+сезон)",
    "weather_snow_model_strength": "Сила зимней snow-модели",
    "weather_snow_export_phys_amp": "Snow: множитель физики (экспорт)",
    "weather_snow_export_stress_amp": "Snow: множитель stress (экспорт)",
    "weather_snow_export_surface_amp": "Snow: покрытие (экспорт)",
    "weather_heat_continuous": "Непрерывная тепло-модель",
    "heat_tree_shade_bonus": "Heat: бонус тени деревьев",
    "heat_open_sky_penalty": "Heat: штраф открытого неба",
    "heat_building_shade_bonus": "Heat: бонус тени зданий",
    "heat_covered_bonus": "Heat: бонус укрытий",
    "heat_wind_open_penalty": "Heat: ветровой штраф (открыто)",
    "heat_wet_surface_penalty": "Heat: мокрое покрытие",
    "heat_norm_temp": "Heat: норм. температура",
    "heat_norm_rain": "Heat: норм. осадки",
    "heat_norm_wind": "Heat: норм. ветер",
    "heat_norm_gust": "Heat: норм. порывы",
    "heat_norm_cloud": "Heat: норм. облачность",
    "heat_norm_humidity": "Heat: норм. влажность",
    "heat_norm_cold_like": "Heat: норм. прохлада",
    "heat_norm_snow_depth": "Heat: норм. глубина снега",
    "heat_norm_snow_fresh": "Heat: норм. снегопад",
    "route_open_sky_share": "Маршрут: доля открытого неба",
    "route_building_shade_share": "Маршрут: тень зданий",
    "route_covered_share": "Маршрут: укрытия",
    "route_bad_wet_surface_share": "Маршрут: плохое мокрое покрытие",
    "route_winter_harsh_surface_share": "Маршрут: тяжёлое зимнее покрытие (прокси)",
    "route_wind_direction_aware": "Маршрут: учтено направление ветра (0/1)",
    "route_mean_wind_to_street_angle_deg": "Маршрут: средний угол улица–ветер, °",
    "route_mean_heat_directional_wind_exp": "Маршрут: средняя ветроэкспозиция (heat)",
    "route_mean_heat_building_wind_factor": "Маршрут: средний фактор экрана зданий",
    "route_frac_wind_along_open_hostile": "Маршрут: доля «вдоль ветра и открыто»",
    "route_frac_wind_cross_building_screen": "Маршрут: доля «поперёк и экран зданий»",
    "route_mean_base_route_factor": "Маршрут: ср. base_route_factor",
    "route_mean_slope_weather_factor": "Маршрут: ср. slope_weather_factor",
    "route_mean_surface_weather_factor": "Маршрут: ср. surface_weather_factor",
    "route_mean_green_route_factor": "Маршрут: ср. green_route_factor",
    "route_mean_open_sky_weather_factor": "Маршрут: ср. open_sky_weather_factor",
    "route_mean_building_shelter_factor": "Маршрут: ср. building_shelter_factor",
    "route_mean_covered_shelter_factor": "Маршрут: ср. covered_shelter_factor",
    "route_mean_stress_weather_factor": "Маршрут: ср. stress_weather_factor (ребро)",
    "route_mean_stairs_weather_factor": "Маршрут: ср. stairs_weather_factor",
    "route_mean_wind_orientation_factor": "Маршрут: ср. wind_orientation_factor",
    "stress_route_regime_factor": "Маршрут: глобальный stress-regime (как в весах)",
    "route_stairs_length_m": "Маршрут: длина по лестницам, м",
    "route_stairs_length_fraction": "Маршрут: доля длины по лестницам",
    "route_winter_open_stress_proxy": "Маршрут: открытость × сила зимней модели",
    "weather_test_case_id": "Test: ID сценария",
    "weather_test_temperature_c": "Test: температура, °C",
    "weather_test_precipitation_mm": "Test: осадки, мм/ч",
    "weather_test_wind_speed_ms": "Test: ветер, м/с",
    "weather_test_wind_gusts_ms": "Test: порывы, м/с",
    "weather_test_wind_direction_deg": "Test: направление ветра, ° (откуда)",
    "weather_test_cloud_cover_pct": "Test: облачность, %",
    "weather_test_humidity_pct": "Test: влажность, %",
    "weather_test_shortwave_radiation_wm2": "Test: КВ, Вт/м²",
    "weather_test_snowfall_cm_h": "Test: снегопад, см/ч",
    "weather_test_snow_depth_m": "Test: глубина снега, м",
    "weather_test_weather_code": "Test: код погоды WMO",
    "weather_test_time_iso": "Test: ISO времени сценария",
    "length_m": "Длина, м",
    "length_km": "Длина, км",
    "time_s": "Время, с",
    "time_min": "Время, мин",
    "time_display": "Время (текст)",
    "climb_m": "Набор, м",
    "descent_m": "Спуск, м",
    "max_gradient_pct": "Макс. уклон, %",
    "avg_gradient_pct": "Средн. уклон, %",
    "max_above_start_m": "Макс. выше старта, м",
    "max_below_start_m": "Макс. ниже старта, м",
    "end_diff_m": "Перепад финиша, м",
    "green_percent": "Озеленение, %",
    "avg_trees_pct": "Деревья, %",
    "avg_grass_pct": "Трава, %",
    "stress_cost_total": "Стресс, суммарно",
    "avg_stress_lts": "Средний LTS",
    "max_stress_lts": "Макс. LTS",
    "high_stress_segments_count": "Высокий стресс, сегм.",
    "stressful_intersections_count": "Стресс. пересечения",
    "cost": "Стоимость модели",
    "mode": "Режим (техн.)",
    "warnings_count": "Число предупреждений",
    "warnings_text": "Предупреждения",
    "geometry_json": "Геометрия (JSON)",
    "full_baseline_length_m": "Длина эталона full, м",
    "delta_length_vs_full_m": "Δ длины к full, м",
    "delta_climb_vs_full_m": "Δ набора к full, м",
}


def _rename_row_keys_ru(row: Dict[str, Any]) -> Dict[str, Any]:
    return {_ROUTE_COL_RU.get(k, k): v for k, v in row.items()}


def write_route_vertices_csv_gzip(path: str, vertices: List[Dict[str, Any]]) -> None:
    import csv
    import gzip

    if vertices and "weather_date" in vertices[0]:
        fields = ["weather_date", "route_id", "vertex_index", "lat", "lon"]
    else:
        fields = ["route_id", "vertex_index", "lat", "lon"]

    with gzip.open(path, "wt", encoding="utf-8", newline="") as gz:
        w = csv.DictWriter(gz, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for row in vertices:
            w.writerow(row)
    _log.info("Вершины маршрутов: %s (%d строк)", path, len(vertices))


def write_xlsx(
    path: str,
    *,
    meta_rows: List[Tuple[str, Any]],
    points: List[SampledPoint],
    raw_rows: List[Dict[str, Any]],
    vertices: List[Dict[str, Any]],
    failures: List[Dict[str, Any]],
    summary_variant: List[Dict[str, Any]],
    summary_profile: List[Dict[str, Any]],
    summary_direction: List[Dict[str, Any]],
    pair_cmp: List[Dict[str, Any]],
    legacy_multisheet: bool = False,
    summary_by_weather_date: Optional[List[Dict[str, Any]]] = None,
    summary_by_weather_date_variant: Optional[List[Dict[str, Any]]] = None,
    summary_heat_kpi: Optional[List[Dict[str, Any]]] = None,
    summary_winter_kpi: Optional[List[Dict[str, Any]]] = None,
    summary_extra_blocks: Optional[Sequence[Tuple[str, List[Dict[str, Any]]]]] = None,
) -> None:
    from openpyxl import Workbook

    if legacy_multisheet:
        wb = Workbook()
        wm = wb.active
        wm.title = "Meta"
        wm.cell(row=1, column=1, value="key")
        wm.cell(row=1, column=2, value="value")
        for i, (k, v) in enumerate(meta_rows, start=2):
            wm.cell(row=i, column=1, value=k)
            wm.cell(row=i, column=2, value=v)

        def add_sheet(name: str, headers: List[str], rows: List[Dict[str, Any]]) -> None:
            ws = wb.create_sheet(title=name)
            _write_sheet_table(ws, headers, rows)

        add_sheet(
            "SampledPoints",
            ["point_id", "lat", "lon", "nearest_node_id"],
            [
                {
                    "point_id": _point_id_fmt(sp.idx),
                    "lat": sp.lat,
                    "lon": sp.lon,
                    "nearest_node_id": sp.nearest_node_id,
                }
                for sp in points
            ],
        )

        if raw_rows:
            headers = list(raw_rows[0].keys())
        else:
            headers = []
        add_sheet("RoutesRaw", headers, raw_rows)
        add_sheet(
            "RouteVertices",
            ["route_id", "vertex_index", "lat", "lon"],
            vertices,
        )
        if summary_variant:
            svh = list(summary_variant[0].keys())
        else:
            svh = []
        add_sheet("SummaryByVariant", svh, summary_variant)
        if summary_profile:
            sph = list(summary_profile[0].keys())
        else:
            sph = []
        add_sheet("SummaryByProfile", sph, summary_profile)
        if summary_direction:
            sdh = list(summary_direction[0].keys())
        else:
            sdh = []
        add_sheet("SummaryDirection", sdh, summary_direction)
        if pair_cmp:
            pch = list(pair_cmp[0].keys())
        else:
            pch = []
        add_sheet("PairComparison", pch, pair_cmp)

        fh = [
            "origin_point_id",
            "destination_point_id",
            "profile",
            "variant",
            "error_code",
            "error_message",
        ]
        add_sheet("Failures", fh, failures)

        wb.save(path)
        _log.info("Excel записан (англ. листы): %s", path)
        return

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Метаданные"
    ws0.cell(row=1, column=1, value="Параметр")
    ws0.cell(row=1, column=2, value="Значение")
    for i, (k, v) in enumerate(_meta_rows_ru(meta_rows), start=2):
        ws0.cell(row=i, column=1, value=k)
        ws0.cell(row=i, column=2, value=v)

    ws1 = wb.create_sheet(title="Точки")
    _write_sheet_table(
        ws1,
        ["ID точки", "Широта", "Долгота", "Узел графа"],
        [
            {
                "ID точки": _point_id_fmt(sp.idx),
                "Широта": sp.lat,
                "Долгота": sp.lon,
                "Узел графа": sp.nearest_node_id,
            }
            for sp in points
        ],
    )

    if raw_rows:
        rh = [_ROUTE_COL_RU.get(k, k) for k in raw_rows[0].keys()]
        ru_rows = [_rename_row_keys_ru(dict(rw)) for rw in raw_rows]
    else:
        rh, ru_rows = [], []
    ws2 = wb.create_sheet(title="Маршруты")
    _write_sheet_table(ws2, rh, ru_rows)

    ws3 = wb.create_sheet(title="Сводка")
    rr = 1
    s_wd = summary_by_weather_date or []
    s_wdv = summary_by_weather_date_variant or []
    blocks: List[Tuple[str, List[Dict[str, Any]]]] = []
    if summary_heat_kpi:
        blocks.append(("Heat-weather KPI (synthetic)", summary_heat_kpi))
    if summary_winter_kpi:
        blocks.append(("Winter KPI (synthetic)", summary_winter_kpi))
    if summary_extra_blocks:
        for title, blk in summary_extra_blocks:
            if blk:
                blocks.append((title, list(blk)))
    blocks.extend(
        [
            ("Средние по дате погоды", s_wd),
            ("Средние по дате погоды и варианту", s_wdv),
            ("Средние по варианту", summary_variant),
            ("Средние по профилю", summary_profile),
            ("Средние по направлению", summary_direction),
            ("Сводка по парам O–D", pair_cmp),
        ]
    )
    for title, block in blocks:
        ws3.cell(row=rr, column=1, value=title)
        rr += 1
        if block and block[0]:
            hdr = list(block[0].keys())
            for c, h in enumerate(hdr, start=1):
                ws3.cell(row=rr, column=c, value=h)
            rr += 1
            for i, row in enumerate(block):
                for c, h in enumerate(hdr, start=1):
                    ws3.cell(row=rr + i, column=c, value=row.get(h))
            rr += len(block) + 1
        else:
            rr += 1

    ws4 = wb.create_sheet(title="Ошибки")
    _write_sheet_table(
        ws4,
        [
            "Отправление",
            "Назначение",
            "Профиль",
            "Погодный кейс",
            "Вариант",
            "Код",
            "Сообщение",
        ],
        [
            {
                "Отправление": f.get("origin_point_id"),
                "Назначение": f.get("destination_point_id"),
                "Профиль": f.get("profile"),
                "Погодный кейс": f.get("weather_test_case_id"),
                "Вариант": f.get("variant"),
                "Код": f.get("error_code"),
                "Сообщение": f.get("error_message"),
            }
            for f in failures
        ],
    )

    wb.save(path)
    _log.info("Excel записан (компактный RU): %s", path)


def run_variants_over_weather_cases(*args: Any, **kwargs: Any) -> str:
    """Обёртка: ``bike_router.tools._run_variants_weather_batch.run_variants_over_weather_cases``."""
    from bike_router.tools._run_variants_weather_batch import (
        run_variants_over_weather_cases as _run_impl,
    )

    return _run_impl(*args, **kwargs)
