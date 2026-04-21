"""Пакетный эксперимент по случайным O-D парам внутри PRECACHE_AREA_POLYGON_WKT -> Excel.

Запуск::

    python -m bike_router.tools.route_batch_experiment --weather now

Требования: в .env задан ``PRECACHE_AREA_POLYGON_WKT``, собран precache
(``python -m bike_router.tools.precache_area``), ``GRAPH_CORRIDOR_MODE=true``
без fixed ``AREA_POLYGON_WKT``, чтобы движок использовал граф арены без
пересборки на каждую пару.

Погода (``--weather``): ``none`` — без учёта; ``now`` — Open-Meteo на момент
старта (UTC); ``past`` — архив за N календарных дней (``--past-days``, по
умолчанию 10), слот каждого дня 14:00 локально (``BATCH_WEATHER_TZ`` /
по умолчанию Europe/Samara).

Результат: ``route_experiment_batch_<YYYYMMDD_HHMMSS>.xlsx`` (UTC) и рядом
``*_route_vertices.csv.gz`` при наличии вершин.

Расширение коридора в батче только **10 м → 100 м**. Если маршрута нет и на 100 м,
пара O-D **пропускается** (без попыток 1000/5000 м); в лог пишутся координаты
начала/конца и идентификаторы точек.

Точки O-D отбираются **только** внутри полигона из ``PRECACHE_AREA_POLYGON_WKT``
(арена precache), через ``parse_precache_polygon`` / uniform sampling в bounds
с проверкой ``polygon.contains(point)`` — не из ``AREA_POLYGON_WKT`` и не из
произвольного bbox без привязки к этому WKT.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import sys
import time
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, time as dt_time, timedelta, timezone
from math import asin, cos, radians, sin, sqrt
from typing import Any, Dict, Iterator, List, Optional, Sequence, Tuple

try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None  # type: ignore[misc, assignment]

_log = logging.getLogger(__name__)

# Порядок как в engine._UNIFIED_ROUTE_ORDER
EXPECTED_VARIANTS: Tuple[str, ...] = (
    "shortest",
    "full",
    "green",
    "heat",
    "stress",
    "heat_stress",
)

PROFILES: Tuple[str, ...] = ("cyclist", "pedestrian")

# Перебор буфера коридора в этом эксперименте; дальше — пропуск пары (см. RouteNotFoundError).
EXPERIMENT_CORRIDOR_EXPAND_M: Tuple[float, ...] = (10.0, 100.0)

# Часовой пояс для «локального часа» в multi-day Open-Meteo (можно переопределить BATCH_WEATHER_TZ).
_DEFAULT_BATCH_WEATHER_TZ = "Europe/Samara"

# Фиксированные параметры батча (CLI урезан).
DEFAULT_ROUTE_BATCH_SEED = 42
DEFAULT_MIN_SPACING_M = 100.0
DEFAULT_MAX_SAMPLE_ATTEMPTS = 50_000
# Режим ``past``: столько дней архива, каждый день — 14:00 локально.
PAST_ARCHIVE_DAYS = 10
PAST_ARCHIVE_LOCAL_HOUR = 14


def _batch_profiles_from_arg(arg: str) -> Tuple[str, ...]:
    v = (arg or "both").strip().lower()
    if v == "both":
        return ("cyclist", "pedestrian")
    if v == "cyclist":
        return ("cyclist",)
    if v == "pedestrian":
        return ("pedestrian",)
    raise ValueError("profiles: ожидается both | cyclist | pedestrian")


def _batch_output_xlsx_path() -> str:
    """Уникальное имя: route_experiment_batch_YYYYMMDD_HHMMSS.xlsx (UTC)."""
    return datetime.now(timezone.utc).strftime(
        "route_experiment_batch_%Y%m%d_%H%M%S.xlsx"
    )


def _zoneinfo_or_raise(name: str) -> Any:
    if ZoneInfo is None:
        raise RuntimeError(
            "Нужен Python 3.9+ с tzdata для часового пояса погоды (zoneinfo)."
        )
    return ZoneInfo(name)


def weather_windows_past_local_days(
    *,
    past_days: int,
    local_hour: int,
    tz_name: str,
) -> List[Tuple[str, str]]:
    """Календарные даты в прошлом (вчера, позавчера, …) и ISO-время для Open-Meteo.

    Возвращает список из ``past_days`` пар
    ``(weather_date YYYY-MM-DD, weather_time локальный ISO)``.
    """
    tz = _zoneinfo_or_raise(tz_name)
    if not (0 <= int(local_hour) <= 23):
        raise ValueError("local_hour должен быть 0..23")
    out: List[Tuple[str, str]] = []
    today = datetime.now(tz).date()
    for k in range(1, int(past_days) + 1):
        d = today - timedelta(days=k)
        dt = datetime.combine(d, dt_time(hour=int(local_hour)), tzinfo=tz)
        out.append((d.isoformat(), dt.isoformat()))
    return out


SUMMARY_NUM_KEYS = (
    "length_m",
    "time_s",
    "climb_m",
    "descent_m",
    "max_gradient_pct",
    "green_percent",
    "avg_trees_pct",
    "avg_grass_pct",
    "stress_cost_total",
    "weather_temperature_c",
    "weather_apparent_temperature_c",
    "weather_precipitation_mm",
    "weather_precipitation_probability",
    "weather_wind_speed_ms",
    "weather_wind_gusts_ms",
    "weather_cloud_cover_pct",
    "weather_humidity_pct",
    "weather_shortwave_radiation_wm2",
)


def _weather_metrics_from_route(r: Any) -> Dict[str, Any]:
    """Числовые поля снимка Open-Meteo из ответа движка (для Excel и сводок)."""
    empty: Dict[str, Any] = {
        "weather_temperature_c": None,
        "weather_apparent_temperature_c": None,
        "weather_precipitation_mm": None,
        "weather_precipitation_probability": None,
        "weather_wind_speed_ms": None,
        "weather_wind_gusts_ms": None,
        "weather_cloud_cover_pct": None,
        "weather_humidity_pct": None,
        "weather_shortwave_radiation_wm2": None,
    }
    w = getattr(r, "weather", None)
    if not w or not bool(getattr(w, "enabled", False)):
        return empty
    s = getattr(w, "snapshot", None)
    if s is None:
        return empty

    def _f(name: str) -> Optional[float]:
        v = getattr(s, name, None)
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    return {
        "weather_temperature_c": _f("temperature_c"),
        "weather_apparent_temperature_c": _f("apparent_temperature_c"),
        "weather_precipitation_mm": _f("precipitation_mm"),
        "weather_precipitation_probability": _f("precipitation_probability"),
        "weather_wind_speed_ms": _f("wind_speed_ms"),
        "weather_wind_gusts_ms": _f("wind_gusts_ms"),
        "weather_cloud_cover_pct": _f("cloud_cover_pct"),
        "weather_humidity_pct": _f("humidity_pct"),
        "weather_shortwave_radiation_wm2": _f("shortwave_radiation_wm2"),
    }


def _ensure_pkg_path() -> None:
    root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    if root not in sys.path:
        sys.path.insert(0, root)


def haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Расстояние по поверхности сферы (м)."""
    r = 6371000.0
    p1, p2 = radians(lat1), radians(lat2)
    dphi = radians(lat2 - lat1)
    dl = radians(lon2 - lon1)
    a = sin(dphi / 2) ** 2 + cos(p1) * cos(p2) * sin(dl / 2) ** 2
    return 2 * r * asin(min(1.0, sqrt(a)))


def _point_id_fmt(i: int) -> str:
    return f"P{i + 1:02d}"


@dataclass
class SampledPoint:
    idx: int
    lat: float
    lon: float
    nearest_node_id: int


def _iter_directed_pairs(n: int) -> Iterator[Tuple[int, int]]:
    for i in range(n):
        for j in range(n):
            if i == j:
                continue
            yield i, j


def _direction_key(a: int, b: int) -> str:
    return f"{_point_id_fmt(a)}->{_point_id_fmt(b)}"


def sample_points_in_polygon(
    *,
    poly: Any,
    n_target: int,
    rng: Any,
    engine: Any,
    min_spacing_m: float,
    max_attempts: int,
) -> List[SampledPoint]:
    """Случайные точки в полигоне: уникальность, min_spacing_m, валидация на графе."""
    from shapely.geometry import Point

    import osmnx as ox

    from bike_router.exceptions import PointOutsideZoneError

    minx, miny, maxx, maxy = poly.bounds
    out: List[SampledPoint] = []
    attempts = 0
    while len(out) < n_target and attempts < max_attempts:
        attempts += 1
        x = float(rng.uniform(minx, maxx))
        y = float(rng.uniform(miny, maxy))
        p = Point(x, y)
        if not poly.contains(p):
            continue
        lat, lon = y, x
        if any(
            haversine_m(lat, lon, sp.lat, sp.lon) < min_spacing_m for sp in out
        ):
            continue
        try:
            engine._validate_point((lat, lon), "sample")  # noqa: SLF001
        except PointOutsideZoneError:
            continue
        G = engine.graph
        if G is None:
            raise RuntimeError("Граф не загружен после warmup")
        nid = int(ox.distance.nearest_nodes(G, X=lon, Y=lat))
        out.append(SampledPoint(idx=len(out), lat=lat, lon=lon, nearest_node_id=nid))

    if len(out) < n_target:
        raise RuntimeError(
            f"Не удалось набрать {n_target} точек за {max_attempts} попыток "
            f"(получено {len(out)}). Увеличьте max_attempts или ослабьте min_spacing_m."
        )
    _log.info(
        "Точки: готово %d шт. (попыток %d, min %.0f м)",
        len(out),
        attempts,
        min_spacing_m,
    )
    return out


def _warnings_text(r: Any) -> Tuple[int, str]:
    qh = r.quality_hints
    if not qh or not qh.warnings:
        return 0, ""
    w = qh.warnings
    return len(w), " | ".join(str(x) for x in w)


def _stress_fields(r: Any) -> Dict[str, Any]:
    hs = r.heat_stress
    out: Dict[str, Any] = {
        "stress_cost_total": float(r.stress_cost_total or 0.0),
        "avg_stress_lts": None,
        "max_stress_lts": None,
        "high_stress_segments_count": int(r.high_stress_segments_count or 0),
        "stressful_intersections_count": int(r.stressful_intersections_count or 0),
    }
    if hs:
        out["avg_stress_lts"] = float(hs.avg_stress_lts)
        out["max_stress_lts"] = float(hs.max_stress_lts)
        if out["stress_cost_total"] == 0.0 and hs.stress_cost_total:
            out["stress_cost_total"] = float(hs.stress_cost_total)
        if not out["high_stress_segments_count"] and hs.high_stress_segments_count:
            out["high_stress_segments_count"] = int(hs.high_stress_segments_count)
        if (
            not out["stressful_intersections_count"]
            and hs.stressful_intersections_count
        ):
            out["stressful_intersections_count"] = int(
                hs.stressful_intersections_count
            )
    return out


def route_to_raw_row(
    *,
    experiment_id: str,
    seed: int,
    route_id: int,
    profile: str,
    origin_id: int,
    dest_id: int,
    o_lat: float,
    o_lon: float,
    d_lat: float,
    d_lon: float,
    r: Any,
    baseline_full: Any = None,
    weather_date: str = "",
) -> Dict[str, Any]:
    st = _stress_fields(r)
    wn, wt = _warnings_text(r)
    elev = r.elevation
    green = r.green
    wc = r.weather.weather_time if r.weather else None
    wsrc = str(r.weather.source or "") if r.weather else ""
    geom = r.geometry or []
    geom_json = json.dumps(geom, ensure_ascii=False, separators=(",", ":"))

    out: Dict[str, Any] = {
        "route_id": route_id,
        "experiment_id": experiment_id,
        "seed": seed,
        "profile": profile,
        "variant_key": r.mode,
        "variant_label": r.variant_label or "",
        "origin_point_id": _point_id_fmt(origin_id),
        "destination_point_id": _point_id_fmt(dest_id),
        "direction_key": _direction_key(origin_id, dest_id),
        "direction_order": "forward"
        if origin_id < dest_id
        else "reverse",
        "origin_lat": o_lat,
        "origin_lon": o_lon,
        "destination_lat": d_lat,
        "destination_lon": d_lon,
        "route_built_at_utc": r.route_built_at_utc or "",
        "weather_date": weather_date or (str(wc)[:10] if wc else ""),
        "weather_time": wc or "",
        "weather_source": wsrc,
        **_weather_metrics_from_route(r),
        "length_m": float(r.length_m),
        "length_km": round(float(r.length_m) / 1000.0, 6),
        "time_s": float(r.time_s),
        "time_min": round(float(r.time_s) / 60.0, 4),
        "time_display": r.time_display or "",
        "climb_m": float(elev.climb_m),
        "descent_m": float(elev.descent_m),
        "max_gradient_pct": float(elev.max_gradient_pct),
        "avg_gradient_pct": float(elev.avg_gradient_pct),
        "max_above_start_m": float(elev.max_above_start_m),
        "max_below_start_m": float(elev.max_below_start_m),
        "end_diff_m": float(elev.end_diff_m),
        "green_percent": float(green.percent),
        "avg_trees_pct": float(green.avg_trees_pct),
        "avg_grass_pct": float(green.avg_grass_pct),
        "stress_cost_total": st["stress_cost_total"],
        "avg_stress_lts": st["avg_stress_lts"],
        "max_stress_lts": st["max_stress_lts"],
        "high_stress_segments_count": st["high_stress_segments_count"],
        "stressful_intersections_count": st["stressful_intersections_count"],
        "cost": float(r.cost),
        "mode": r.mode,
        "warnings_count": wn,
        "warnings_text": wt,
        "geometry_json": geom_json,
    }
    if baseline_full is not None:
        out["full_baseline_length_m"] = round(float(baseline_full.length_m), 2)
        be = baseline_full.elevation
        out["delta_length_vs_full_m"] = round(
            float(r.length_m) - float(baseline_full.length_m), 2
        )
        out["delta_climb_vs_full_m"] = round(
            float(elev.climb_m) - float(be.climb_m), 2
        )
    else:
        out["full_baseline_length_m"] = None
        out["delta_length_vs_full_m"] = None
        out["delta_climb_vs_full_m"] = None
    return out


def _mean(xs: List[float]) -> Optional[float]:
    if not xs:
        return None
    return sum(xs) / len(xs)


def _write_sheet_table(
    ws: Any, headers: Sequence[str], rows: List[Dict[str, Any]]
) -> None:
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for ri, row in enumerate(rows, start=2):
        for c, h in enumerate(headers, start=1):
            v = row.get(h)
            ws.cell(row=ri, column=c, value=v)


def _make_summary_row(prefix: Dict[str, Any], rs: List[Dict[str, Any]]) -> Dict[str, Any]:
    row = dict(prefix)
    for k in SUMMARY_NUM_KEYS:
        vals: List[float] = []
        for r in rs:
            v = r.get(k)
            if v is None or v == "":
                continue
            try:
                vals.append(float(v))
            except (TypeError, ValueError):
                continue
        m = _mean(vals)
        row[f"mean_{k}"] = round(m, 4) if m is not None else None
    row["routes_count"] = len(rs)
    return row


def build_summaries(
    raw_rows: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """SummaryByVariant, SummaryByProfile, SummaryDirection."""

    def make_row(prefix: Dict[str, Any], rs: List[Dict[str, Any]]) -> Dict[str, Any]:
        return _make_summary_row(prefix, rs)

    by_pv: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    by_prof: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    by_dir: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for r in raw_rows:
        by_pv[(str(r["profile"]), str(r["variant_key"]))].append(r)
        by_prof[str(r["profile"])].append(r)
        by_dir[str(r["direction_order"])].append(r)

    s_var = [
        make_row({"profile": pk, "variant_key": vk}, rs)
        for (pk, vk), rs in sorted(by_pv.items())
    ]
    s_prof = [make_row({"profile": pk}, rs) for pk, rs in sorted(by_prof.items())]
    s_dir = [
        make_row({"direction_order": dk}, rs) for dk, rs in sorted(by_dir.items())
    ]
    return s_var, s_prof, s_dir


def build_summaries_by_weather_date(
    raw_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Средние по календарной дате погоды (multi-day)."""
    by_d: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        wd = str(r.get("weather_date") or "").strip()
        if wd:
            by_d[wd].append(r)
    return [
        _make_summary_row({"weather_date": d}, rs)
        for d, rs in sorted(by_d.items())
    ]


def build_summaries_by_weather_date_and_variant(
    raw_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Средние по паре (дата погоды, вариант маршрута)."""
    g: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        wd = str(r.get("weather_date") or "").strip()
        vk = str(r.get("variant_key") or "")
        if wd:
            g[(wd, vk)].append(r)
    return [
        _make_summary_row({"weather_date": a, "variant_key": b}, rs)
        for (a, b), rs in sorted(g.items())
    ]


def build_pair_comparison(
    raw_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Сводка по (origin, dest, profile[, дата погоды]): min/max/avg length и time."""
    has_wd = bool(raw_rows and str(raw_rows[0].get("weather_date") or "").strip())
    groups: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
    for r in raw_rows:
        base = (
            r["origin_point_id"],
            r["destination_point_id"],
            str(r["profile"]),
            str(r["direction_key"]),
        )
        key: Tuple[Any, ...] = (
            base + (str(r.get("weather_date") or "").strip(),) if has_wd else base
        )
        groups[key].append(r)

    out: List[Dict[str, Any]] = []
    for key in sorted(groups.keys()):
        rs = groups[key]
        if has_wd:
            o, d, prof, dk, wdate = key
        else:
            o, d, prof, dk = key
            wdate = ""
        lengths = [float(x["length_m"]) for x in rs]
        times = [float(x["time_s"]) for x in rs]
        row: Dict[str, Any] = {
            "origin_point_id": o,
            "destination_point_id": d,
            "profile": prof,
            "direction_key": dk,
            "variants_count": len(rs),
                "length_min_m": min(lengths) if lengths else None,
                "length_max_m": max(lengths) if lengths else None,
                "length_mean_m": _mean(lengths),
                "time_min_s": min(times) if times else None,
                "time_max_s": max(times) if times else None,
                "time_mean_s": _mean(times),
        }
        if has_wd:
            row["weather_date"] = wdate
        out.append(row)
    return out


def _meta_rows_ru(rows: List[Tuple[str, Any]]) -> List[Tuple[str, Any]]:
    ru = {
        "experiment_id": "ID эксперимента",
        "started_at_utc": "Время старта (UTC)",
        "seed": "Seed",
        "n_points": "Число точек",
        "n_directed_pairs": "Направленных пар O–D",
        "batch_profiles": "Профили (батч)",
        "n_profiles": "Профилей",
        "n_variants": "Вариантов маршрута",
        "expected_route_cells": "Ожидалось строк маршрутов",
        "successful_route_rows": "Успешных строк",
        "skipped_od_no_path_buffer_100m": "Пропущено (нет пути до 100 м)",
        "failure_rows": "Строк ошибок",
        "precache_polygon_wkt_sha256": "SHA256 полигона precache",
        "precache_polygon_bounds_lonlat": "Границы полигона (lon/lat)",
        "precache_polygon_wkt_prefix": "Префикс WKT полигона",
        "routing_algo_version": "Версия алгоритма",
        "routing_weights_fingerprint": "Отпечаток весов",
        "elapsed_seconds": "Время работы, с",
        "weather_schedule": "Сценарий погоды (none / now / past)",
        "weather_mode_engine": "Режим погоды (движок)",
        "weather_time_utc_snapshot": "Снимок времени UTC (режим now)",
        "past_days": "Дней архива (режим past)",
        "past_slot_local_hour": "Локальный час слота (режим past)",
        "batch_weather_tz": "Часовой пояс IANA (режим past)",
        "vertices_file": "Файл вершин (отдельно)",
    }
    return [(ru.get(str(k), str(k)), v) for k, v in rows]


_ROUTE_COL_RU: Dict[str, str] = {
    "route_id": "ID маршрута",
    "experiment_id": "ID эксперимента",
    "seed": "Seed",
    "profile": "Профиль",
    "variant_key": "Вариант (ключ)",
    "variant_label": "Вариант",
    "origin_point_id": "Точка отправления",
    "destination_point_id": "Точка назначения",
    "direction_key": "Направление",
    "direction_order": "Порядок",
    "origin_lat": "Широта начала",
    "origin_lon": "Долгота начала",
    "destination_lat": "Широта конца",
    "destination_lon": "Долгота конца",
    "route_built_at_utc": "Построен в (UTC)",
    "weather_date": "Дата погоды",
    "weather_time": "Время погоды (ISO)",
    "weather_source": "Источник погоды",
    "weather_temperature_c": "Температура, °C",
    "weather_apparent_temperature_c": "Ощущается, °C",
    "weather_precipitation_mm": "Осадки, мм/ч",
    "weather_precipitation_probability": "Вероятность осадков, %",
    "weather_wind_speed_ms": "Ветер, м/с",
    "weather_wind_gusts_ms": "Порывы ветра, м/с",
    "weather_cloud_cover_pct": "Облачность, %",
    "weather_humidity_pct": "Влажность, %",
    "weather_shortwave_radiation_wm2": "КВ радиация, Вт/м²",
    "length_m": "Длина, м",
    "length_km": "Длина, км",
    "time_s": "Время, с",
    "time_min": "Время, мин",
    "time_display": "Время (текст)",
    "climb_m": "Набор, м",
    "descent_m": "Спуск, м",
    "max_gradient_pct": "Макс. уклон, %",
    "avg_gradient_pct": "Средн. уклон, %",
    "max_above_start_m": "Макс. выше старта, м",
    "max_below_start_m": "Макс. ниже старта, м",
    "end_diff_m": "Перепад финиша, м",
    "green_percent": "Озеленение, %",
    "avg_trees_pct": "Деревья, %",
    "avg_grass_pct": "Трава, %",
    "stress_cost_total": "Стресс, суммарно",
    "avg_stress_lts": "Средний LTS",
    "max_stress_lts": "Макс. LTS",
    "high_stress_segments_count": "Высокий стресс, сегм.",
    "stressful_intersections_count": "Стресс. пересечения",
    "cost": "Стоимость модели",
    "mode": "Режим (техн.)",
    "warnings_count": "Число предупреждений",
    "warnings_text": "Предупреждения",
    "geometry_json": "Геометрия (JSON)",
    "full_baseline_length_m": "Длина эталона full, м",
    "delta_length_vs_full_m": "Δ длины к full, м",
    "delta_climb_vs_full_m": "Δ набора к full, м",
}


def _rename_row_keys_ru(row: Dict[str, Any]) -> Dict[str, Any]:
    return {_ROUTE_COL_RU.get(k, k): v for k, v in row.items()}


def write_route_vertices_csv_gzip(path: str, vertices: List[Dict[str, Any]]) -> None:
    import csv
    import gzip

    if vertices and "weather_date" in vertices[0]:
        fields = ["weather_date", "route_id", "vertex_index", "lat", "lon"]
    else:
        fields = ["route_id", "vertex_index", "lat", "lon"]

    with gzip.open(path, "wt", encoding="utf-8", newline="") as gz:
        w = csv.DictWriter(gz, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for row in vertices:
            w.writerow(row)
    _log.info("Вершины маршрутов: %s (%d строк)", path, len(vertices))


def write_xlsx(
    path: str,
    *,
    meta_rows: List[Tuple[str, Any]],
    points: List[SampledPoint],
    raw_rows: List[Dict[str, Any]],
    vertices: List[Dict[str, Any]],
    failures: List[Dict[str, Any]],
    summary_variant: List[Dict[str, Any]],
    summary_profile: List[Dict[str, Any]],
    summary_direction: List[Dict[str, Any]],
    pair_cmp: List[Dict[str, Any]],
    legacy_multisheet: bool = False,
    summary_by_weather_date: Optional[List[Dict[str, Any]]] = None,
    summary_by_weather_date_variant: Optional[List[Dict[str, Any]]] = None,
) -> None:
    from openpyxl import Workbook

    if legacy_multisheet:
        wb = Workbook()
        wm = wb.active
        wm.title = "Meta"
        wm.cell(row=1, column=1, value="key")
        wm.cell(row=1, column=2, value="value")
        for i, (k, v) in enumerate(meta_rows, start=2):
            wm.cell(row=i, column=1, value=k)
            wm.cell(row=i, column=2, value=v)

        def add_sheet(name: str, headers: List[str], rows: List[Dict[str, Any]]) -> None:
            ws = wb.create_sheet(title=name)
            _write_sheet_table(ws, headers, rows)

        add_sheet(
            "SampledPoints",
            ["point_id", "lat", "lon", "nearest_node_id"],
            [
                {
                    "point_id": _point_id_fmt(sp.idx),
                    "lat": sp.lat,
                    "lon": sp.lon,
                    "nearest_node_id": sp.nearest_node_id,
                }
                for sp in points
            ],
        )

        if raw_rows:
            headers = list(raw_rows[0].keys())
        else:
            headers = []
        add_sheet("RoutesRaw", headers, raw_rows)
        add_sheet(
            "RouteVertices",
            ["route_id", "vertex_index", "lat", "lon"],
            vertices,
        )
        if summary_variant:
            svh = list(summary_variant[0].keys())
        else:
            svh = []
        add_sheet("SummaryByVariant", svh, summary_variant)
        if summary_profile:
            sph = list(summary_profile[0].keys())
        else:
            sph = []
        add_sheet("SummaryByProfile", sph, summary_profile)
        if summary_direction:
            sdh = list(summary_direction[0].keys())
        else:
            sdh = []
        add_sheet("SummaryDirection", sdh, summary_direction)
        if pair_cmp:
            pch = list(pair_cmp[0].keys())
        else:
            pch = []
        add_sheet("PairComparison", pch, pair_cmp)

        fh = [
            "origin_point_id",
            "destination_point_id",
            "profile",
            "variant",
            "error_code",
            "error_message",
        ]
        add_sheet("Failures", fh, failures)

        wb.save(path)
        _log.info("Excel записан (англ. листы): %s", path)
        return

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Метаданные"
    ws0.cell(row=1, column=1, value="Параметр")
    ws0.cell(row=1, column=2, value="Значение")
    for i, (k, v) in enumerate(_meta_rows_ru(meta_rows), start=2):
        ws0.cell(row=i, column=1, value=k)
        ws0.cell(row=i, column=2, value=v)

    ws1 = wb.create_sheet(title="Точки")
    _write_sheet_table(
        ws1,
        ["ID точки", "Широта", "Долгота", "Узел графа"],
        [
            {
                "ID точки": _point_id_fmt(sp.idx),
                "Широта": sp.lat,
                "Долгота": sp.lon,
                "Узел графа": sp.nearest_node_id,
            }
            for sp in points
        ],
    )

    if raw_rows:
        rh = [_ROUTE_COL_RU.get(k, k) for k in raw_rows[0].keys()]
        ru_rows = [_rename_row_keys_ru(dict(rw)) for rw in raw_rows]
    else:
        rh, ru_rows = [], []
    ws2 = wb.create_sheet(title="Маршруты")
    _write_sheet_table(ws2, rh, ru_rows)

    ws3 = wb.create_sheet(title="Сводка")
    rr = 1
    s_wd = summary_by_weather_date or []
    s_wdv = summary_by_weather_date_variant or []
    for title, block in (
        ("Средние по дате погоды", s_wd),
        ("Средние по дате погоды и варианту", s_wdv),
        ("Средние по варианту", summary_variant),
        ("Средние по профилю", summary_profile),
        ("Средние по направлению", summary_direction),
        ("Сводка по парам O–D", pair_cmp),
    ):
        ws3.cell(row=rr, column=1, value=title)
        rr += 1
        if block and block[0]:
            hdr = list(block[0].keys())
            for c, h in enumerate(hdr, start=1):
                ws3.cell(row=rr, column=c, value=h)
            rr += 1
            for i, row in enumerate(block):
                for c, h in enumerate(hdr, start=1):
                    ws3.cell(row=rr + i, column=c, value=row.get(h))
            rr += len(block) + 1
        else:
            rr += 1

    ws4 = wb.create_sheet(title="Ошибки")
    _write_sheet_table(
        ws4,
        ["Отправление", "Назначение", "Профиль", "Вариант", "Код", "Сообщение"],
        [
            {
                "Отправление": f.get("origin_point_id"),
                "Назначение": f.get("destination_point_id"),
                "Профиль": f.get("profile"),
                "Вариант": f.get("variant"),
                "Код": f.get("error_code"),
                "Сообщение": f.get("error_message"),
            }
            for f in failures
        ],
    )

    wb.save(path)
    _log.info("Excel записан (компактный RU): %s", path)


def run_experiment(
    *,
    n_points: int = 10,
    seed: int = DEFAULT_ROUTE_BATCH_SEED,
    min_spacing_m: float = DEFAULT_MIN_SPACING_M,
    max_sample_attempts: int = DEFAULT_MAX_SAMPLE_ATTEMPTS,
    weather: str = "none",
    past_archive_days: int = PAST_ARCHIVE_DAYS,
    precipitation_mm: Optional[float] = None,
    profiles_mode: str = "both",
) -> str:
    _ensure_pkg_path()

    from bike_router.config import ROUTING_ALGO_VERSION, Settings, routing_engine_cache_fingerprint
    from bike_router.engine import RouteEngine
    from bike_router.exceptions import BikeRouterError, RouteNotFoundError
    from bike_router.services.area_graph_cache import parse_precache_polygon

    import numpy as np

    settings = Settings()
    _log.info("Каталог данных: base_dir=%s (cache=%s)", settings.base_dir, settings.cache_dir)
    if not settings.has_precache_area_polygon:
        raise SystemExit(
            "В .env должен быть задан PRECACHE_AREA_POLYGON_WKT (полигон арены)."
        )

    experiment_id = str(uuid.uuid4())
    wkt_stripped = settings.precache_area_polygon_wkt_stripped
    precache_wkt_sha256 = hashlib.sha256(
        wkt_stripped.encode("utf-8")
    ).hexdigest()
    poly = parse_precache_polygon(settings)
    minx, miny, maxx, maxy = poly.bounds
    rng = np.random.default_rng(seed)

    t_start = time.perf_counter()
    _log.info(
        "Арена выборки точек: PRECACHE_AREA_POLYGON_WKT (sha256=%s…, bounds lon[%.5f..%.5f] lat[%.5f..%.5f])",
        precache_wkt_sha256[:16],
        minx,
        maxx,
        miny,
        maxy,
    )
    _log.info("Warmup движка (area_precache / граф)…")
    eng = RouteEngine()
    eng.warmup()

    if eng.graph is None or not eng.is_loaded:
        raise SystemExit("Warmup не загрузил граф. Проверьте .env и precache_area.")

    _log.info("Генерация %d точек (seed=%s)…", n_points, seed)
    points = sample_points_in_polygon(
        poly=poly,
        n_target=n_points,
        rng=rng,
        engine=eng,
        min_spacing_m=min_spacing_m,
        max_attempts=max_sample_attempts,
    )
    n_pairs = n_points * (n_points - 1)
    tz_resolved = os.getenv("BATCH_WEATHER_TZ", _DEFAULT_BATCH_WEATHER_TZ)

    sched = (weather or "none").strip().lower()
    if sched not in ("none", "now", "past"):
        raise ValueError("weather должен быть one of: none, now, past")

    past_days = 0
    local_hour = PAST_ARCHIVE_LOCAL_HOUR
    weather_time_snapshot: Optional[str] = None

    if sched == "past":
        pd = int(past_archive_days)
        if pd < 1:
            raise ValueError(
                "past_archive_days (параметр --past-days) для режима past должен быть >= 1"
            )
        past_days = pd
        weather_windows = weather_windows_past_local_days(
            past_days=past_days,
            local_hour=local_hour,
            tz_name=tz_resolved,
        )
        wm_engine = "auto"
        use_live_engine = True
        _log.info(
            "Погода past: %d дней, локальный час %02d:00 (%s), первый слот %s",
            len(weather_windows),
            local_hour,
            tz_resolved,
            weather_windows[0][0] if weather_windows else "—",
        )
    elif sched == "now":
        weather_time_snapshot = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        weather_windows = [("", weather_time_snapshot)]
        wm_engine = "auto"
        use_live_engine = True
        _log.info("Погода now: снимок UTC %s", weather_time_snapshot)
    else:
        weather_windows = [("", None)]
        wm_engine = "none"
        use_live_engine = False

    profile_tuple = _batch_profiles_from_arg(profiles_mode)
    n_win = max(1, len(weather_windows))
    expected_routes = n_pairs * len(profile_tuple) * len(EXPECTED_VARIANTS) * n_win

    raw_rows: List[Dict[str, Any]] = []
    vertices: List[Dict[str, Any]] = []
    failures: List[Dict[str, Any]] = []

    route_id_counter = 1
    n_ok_route_rows = 0
    n_fail_rows = 0
    n_skipped_no_path_100m = 0

    route_tasks: List[Tuple[int, int, str]] = [
        (i, j, prof)
        for i, j in _iter_directed_pairs(n_points)
        for prof in profile_tuple
    ]

    total_steps = len(route_tasks) * n_win
    from tqdm import tqdm

    pbar: Any = tqdm(
        total=total_steps,
        desc="Маршруты",
        unit="запрос",
        leave=True,
        mininterval=0.3,
    )

    for wdate_str, wt_iso in weather_windows:
        for i, j, prof in route_tasks:
            o_pt = points[i]
            d_pt = points[j]
            start = (o_pt.lat, o_pt.lon)
            end = (d_pt.lat, d_pt.lon)
            pbar.set_postfix(
                od=_direction_key(i, j),
                day=wdate_str or "—",
                p=prof,
                ok=n_ok_route_rows,
                fail=n_fail_rows,
                skip=n_skipped_no_path_100m,
                refresh=False,
            )
            by_mode: Dict[str, Any] = {}
            try:
                alt = eng.compute_alternatives(
                    start,
                    end,
                    prof,
                    green_enabled=True,
                    weather_mode=wm_engine,
                    use_live_weather=use_live_engine,
                    weather_time=wt_iso if wt_iso else weather_time_snapshot,
                    precipitation_mm=precipitation_mm,
                    corridor_expand_schedule_meters=EXPERIMENT_CORRIDOR_EXPAND_M,
                )
                by_mode = {r.mode: r for r in alt.routes}
            except RouteNotFoundError as e:
                n_skipped_no_path_100m += 1
                _log.warning(
                    "Пропуск O-D: нет маршрута при буфере коридора до 100 м. "
                    "day=%s origin lat=%.6f lon=%.6f | dest lat=%.6f lon=%.6f | profile=%s | "
                    "points %s→%s (%s). %s",
                    wdate_str or "—",
                    start[0],
                    start[1],
                    end[0],
                    end[1],
                    prof,
                    _point_id_fmt(i),
                    _point_id_fmt(j),
                    _direction_key(i, j),
                    e,
                )
                pbar.update(1)
                continue
            except (BikeRouterError, Exception) as e:
                code = getattr(e, "code", type(e).__name__)
                for mode in EXPECTED_VARIANTS:
                    fl = {
                        "origin_point_id": _point_id_fmt(i),
                        "destination_point_id": _point_id_fmt(j),
                        "profile": prof,
                        "variant": mode,
                        "error_code": str(code),
                        "error_message": str(e),
                    }
                    if wdate_str:
                        fl["weather_date"] = wdate_str
                    failures.append(fl)
                n_fail_rows += 1
                pbar.update(1)
                continue

            for mode in EXPECTED_VARIANTS:
                if mode not in by_mode:
                    fl = {
                        "origin_point_id": _point_id_fmt(i),
                        "destination_point_id": _point_id_fmt(j),
                        "profile": prof,
                        "variant": mode,
                        "error_code": "MISSING_VARIANT",
                        "error_message": "Маршрут не включён в ответ движка",
                    }
                    if wdate_str:
                        fl["weather_date"] = wdate_str
                    failures.append(fl)
                    n_fail_rows += 1
                    continue

                r = by_mode[mode]
                rid = route_id_counter
                route_id_counter += 1
                row = route_to_raw_row(
                    experiment_id=experiment_id,
                    seed=seed,
                    route_id=rid,
                    profile=prof,
                    origin_id=i,
                    dest_id=j,
                    o_lat=o_pt.lat,
                    o_lon=o_pt.lon,
                    d_lat=d_pt.lat,
                    d_lon=d_pt.lon,
                    r=r,
                    baseline_full=by_mode.get("full"),
                    weather_date=wdate_str,
                )
                raw_rows.append(row)
                n_ok_route_rows += 1
                for vi, pt in enumerate(r.geometry or []):
                    if len(pt) >= 2:
                        vx: Dict[str, Any] = {
                            "route_id": rid,
                            "vertex_index": vi,
                            "lat": float(pt[0]),
                            "lon": float(pt[1]),
                        }
                        if wdate_str:
                            vx["weather_date"] = wdate_str
                        vertices.append(vx)
            pbar.update(1)

    pbar.close()

    s_var, s_prof, s_dir = build_summaries(raw_rows)
    s_wd = build_summaries_by_weather_date(raw_rows)
    s_wdv = build_summaries_by_weather_date_and_variant(raw_rows)
    pair_cmp = build_pair_comparison(raw_rows)

    out = _batch_output_xlsx_path()

    wkt_fp = routing_engine_cache_fingerprint()
    meta_rows: List[Tuple[str, Any]] = [
        ("experiment_id", experiment_id),
        ("started_at_utc", datetime.now(timezone.utc).isoformat()),
        ("seed", seed),
        ("n_points", n_points),
        ("n_directed_pairs", n_pairs),
        ("batch_profiles", ",".join(profile_tuple)),
        ("n_profiles", len(profile_tuple)),
        ("n_variants", len(EXPECTED_VARIANTS)),
        ("expected_route_cells", expected_routes),
        ("successful_route_rows", len(raw_rows)),
        ("skipped_od_no_path_buffer_100m", n_skipped_no_path_100m),
        ("failure_rows", len(failures)),
        ("precache_polygon_wkt_sha256", precache_wkt_sha256),
        (
            "precache_polygon_bounds_lonlat",
            f"{minx:.8f},{miny:.8f},{maxx:.8f},{maxy:.8f}",
        ),
        ("precache_polygon_wkt_prefix", settings.precache_area_polygon_wkt_stripped[:120]),
        ("routing_algo_version", ROUTING_ALGO_VERSION),
        ("routing_weights_fingerprint", wkt_fp),
        ("weather_schedule", sched),
        ("weather_mode_engine", wm_engine),
        ("weather_time_utc_snapshot", weather_time_snapshot or ""),
        ("past_days", int(past_days)),
        ("past_slot_local_hour", local_hour if int(past_days) > 0 else ""),
        ("batch_weather_tz", tz_resolved if int(past_days) > 0 else ""),
        ("elapsed_seconds", round(time.perf_counter() - t_start, 2)),
    ]

    vgz_path: Optional[str] = None
    if vertices:
        stem = out[:-5] if out.lower().endswith(".xlsx") else out
        vgz_path = f"{stem}_route_vertices.csv.gz"
        write_route_vertices_csv_gzip(vgz_path, vertices)
        meta_rows = meta_rows + [("vertices_file", vgz_path)]

    if n_skipped_no_path_100m:
        _log.info(
            "Пропущено направлений O-D (нет пути после буфера 100 м): %d",
            n_skipped_no_path_100m,
        )
    _log.info("Запись Excel (%d строк маршрутов, вершин %d)…", len(raw_rows), len(vertices))
    write_xlsx(
        out,
        meta_rows=meta_rows,
        points=points,
        raw_rows=raw_rows,
        vertices=vertices,
        failures=failures,
        summary_variant=s_var,
        summary_profile=s_prof,
        summary_direction=s_dir,
        pair_cmp=pair_cmp,
        legacy_multisheet=False,
        summary_by_weather_date=s_wd,
        summary_by_weather_date_variant=s_wdv,
    )
    _log.info("Готово за %.1f с.", time.perf_counter() - t_start)
    return out


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Пакетный эксперимент O-D в PRECACHE_AREA_POLYGON_WKT -> "
            "route_experiment_batch_<YYYYMMDD_HHMMSS>.xlsx (UTC) и вершины *_route_vertices.csv.gz."
        )
    )
    parser.add_argument(
        "--n-points",
        type=int,
        default=10,
        help="Число случайных точек",
    )
    parser.add_argument(
        "--weather",
        choices=("none", "now", "past"),
        default="none",
        help=(
            "none — без учёта погоды; "
            "now — Open-Meteo на дату/время старта эксперимента (UTC); "
            "past — архив за N календарных дней (см. --past-days), каждый день "
            f"{PAST_ARCHIVE_LOCAL_HOUR:02d}:00 локально (BATCH_WEATHER_TZ или Europe/Samara)."
        ),
    )
    parser.add_argument(
        "--past-days",
        type=int,
        default=PAST_ARCHIVE_DAYS,
        metavar="N",
        help=(
            "Для --weather past: сколько календарных дней архива (от вчера назад). "
            f"По умолчанию {PAST_ARCHIVE_DAYS}. Для none/now не используется."
        ),
    )
    parser.add_argument(
        "--precipitation-mm",
        type=float,
        default=None,
        help="Осадки, мм/ч",
    )
    parser.add_argument(
        "--profiles",
        choices=("both", "cyclist", "pedestrian"),
        default="both",
        help="Профили: только велосипед, только пешеход или оба (по умолчанию).",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )

    path = run_experiment(
        n_points=args.n_points,
        weather=args.weather,
        past_archive_days=args.past_days,
        precipitation_mm=args.precipitation_mm,
        profiles_mode=args.profiles,
    )
    print(path)


if __name__ == "__main__":
    main()
